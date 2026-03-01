#!/usr/bin/env python3
"""
Nigeria PolSim: 2026 -> 2058 LGA Data Projection Script
Reads nigeria_lga_polsim_formatted_fixed.xlsx (774 LGAs, 127 columns)
Produces nigeria_lga_polsim_2058.xlsx with 32 years of cumulative transformation.
"""

import re
import math
import copy
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

# ============================================================================
# SECTION 1: CONSTANTS AND CONFIGURATION
# ============================================================================

INPUT_FILE = "nigeria_lga_polsim_formatted_fixed.xlsx"
OUTPUT_FILE = "nigeria_lga_polsim_2058.xlsx"
SEED = 42
np.random.seed(SEED)

START_YEAR = 2026
END_YEAR = 2058

# Column indices (0-based) in the original data
COL = {
    'state': 0, 'lga_name': 1, 'colonial_region': 2, 'terrain': 3,
    'population': 4, 'pop_density': 5, 'median_age': 6, 'pct_under_30': 7,
    # Ethnic columns: 8 through 93 (86 groups)
    'ethnic_start': 8, 'ethnic_end': 93,
    'pct_pada': 91, 'pct_naijin': 92, 'pct_other': 93,
    'pct_muslim': 94, 'pct_christian': 95, 'pct_traditionalist': 96,
    'religious_subtype_notes': 97, 'almajiri_prevalence': 98,
    'major_urban_center': 99, 'urban_rural_split': 100,
    'oil_producing': 101, 'dominant_livelihood': 102,
    'resource_extraction_notes': 103,
    'poverty_rate': 104, 'unemployment_rate': 105, 'youth_unemployment': 106,
    'access_electricity': 107, 'access_water': 108, 'access_healthcare': 109,
    'road_quality': 110, 'market_access': 111,
    'adult_literacy': 112, 'male_literacy': 113, 'female_literacy': 114,
    'primary_enrollment': 115, 'secondary_enrollment': 116,
    'gender_parity': 117, 'out_of_school': 118,
    'num_secondary_schools': 119, 'tertiary_institution': 120,
    'trad_authority': 121, 'trad_authority_influence': 122,
    'predominant_land_tenure': 123,
    'mobile_penetration': 124, 'internet_access': 125,
    'data_notes': 126,
}

# Administrative Zone assignments: state -> (AZ number, AZ name)
AZ_MAP = {
    'Lagos': (1, 'Federal Capital Zone'),
    'Ogun': (2, 'Niger Zone'), 'Oyo': (2, 'Niger Zone'),
    'Kwara': (2, 'Niger Zone'), 'Niger': (2, 'Niger Zone'),
    'Osun': (3, 'Confluence Zone'), 'Ekiti': (3, 'Confluence Zone'),
    'Ondo': (3, 'Confluence Zone'), 'Edo': (3, 'Confluence Zone'),
    'Kogi': (3, 'Confluence Zone'),
    'Delta': (4, 'Littoral Zone'), 'Bayelsa': (4, 'Littoral Zone'),
    'Rivers': (4, 'Littoral Zone'), 'Akwa Ibom': (4, 'Littoral Zone'),
    'Cross River': (4, 'Littoral Zone'),
    'Abia': (5, 'Eastern Zone'), 'Imo': (5, 'Eastern Zone'),
    'Anambra': (5, 'Eastern Zone'), 'Enugu': (5, 'Eastern Zone'),
    'Ebonyi': (5, 'Eastern Zone'), 'Benue': (5, 'Eastern Zone'),
    'Kano': (6, 'Central Zone'), 'Plateau': (6, 'Central Zone'),
    'Nasarawa': (6, 'Central Zone'), 'FCT': (6, 'Central Zone'),
    'Borno': (7, 'Chad Zone'), 'Yobe': (7, 'Chad Zone'),
    'Gombe': (7, 'Chad Zone'), 'Bauchi': (7, 'Chad Zone'),
    'Jigawa': (7, 'Chad Zone'), 'Adamawa': (7, 'Chad Zone'),
    'Taraba': (7, 'Chad Zone'),
    'Kaduna': (8, 'Savanna Zone'), 'Katsina': (8, 'Savanna Zone'),
    'Zamfara': (8, 'Savanna Zone'), 'Sokoto': (8, 'Savanna Zone'),
    'Kebbi': (8, 'Savanna Zone'),
}

NORTHERN_STATES = {
    'Adamawa', 'Bauchi', 'Benue', 'Borno', 'FCT', 'Gombe', 'Jigawa',
    'Kaduna', 'Kano', 'Katsina', 'Kebbi', 'Kogi', 'Kwara', 'Nasarawa',
    'Niger', 'Plateau', 'Sokoto', 'Taraba', 'Yobe', 'Zamfara'
}

SOUTHERN_STATES = {
    'Abia', 'Akwa Ibom', 'Anambra', 'Bayelsa', 'Cross River', 'Delta',
    'Ebonyi', 'Edo', 'Ekiti', 'Enugu', 'Imo', 'Lagos', 'Ogun', 'Ondo',
    'Osun', 'Oyo', 'Rivers'
}

# GDP per capita trajectory (year -> USD)
GDP_TRAJECTORY = {
    2026: 800, 2032: 900, 2037: 2000, 2042: 3700,
    2047: 7000, 2052: 13000, 2057: 26000, 2058: 28000
}

# Rail corridor: primary route states (Lagos-Gausa high-speed)
PRIMARY_RAIL_STATES = ['Lagos', 'Ogun', 'Oyo', 'Kwara', 'Niger', 'FCT', 'Kaduna', 'Kano', 'Katsina']

# Secondary rail branches and approximate start years
SECONDARY_RAIL = {
    2034: [('Kaduna', 'Sokoto')],  # Northern expansion
    2035: [('Kano', 'Jigawa'), ('Kaduna', 'Zamfara')],
    2037: [('Lagos', 'Ogun', 'Ondo', 'Edo', 'Delta', 'Anambra'),  # SE corridor
           ('Rivers', 'Abia', 'Enugu')],  # PH-Enugu
    2038: [('Kano', 'Bauchi', 'Gombe', 'Borno')],  # Kano-Maiduguri
    2039: [('FCT', 'Nasarawa', 'Plateau')],  # Abuja-Jos
    2042: [('Kaduna', 'Kebbi')],  # Kaduna-Sokoto extension
    2045: [('Kogi', 'Benue')],  # Feeder to refinery zone
    2048: [('Borno', 'Yobe'), ('Taraba', 'Adamawa')],  # Far northeast
}

# Planned cities: (name, state, LGA_name_substring, start_year)
PLANNED_CITIES = [
    ('New Anka', 'Zamfara', 'Anka', 2035),
    ('New Makoko', 'Lagos', 'Ibeju', 2037),
    ('Eko Atlantic', 'Lagos', 'Eti-Osa', 2032),  # completed 2039
    ('New Onitsha', 'Anambra', 'Onitsha', 2039),
    ('Abuja South', 'Nasarawa', 'Karu', 2041),
    ('Bonny Hub', 'Rivers', 'Bonny', 2043),
    ('Benue Agri-City', 'Benue', 'Makurdi', 2045),
    ('Plateau New Town', 'Plateau', 'Jos South', 2048),
]

# Refinery zones: (state, LGA substrings, operational year)
REFINERIES = [
    ('Rivers', ['Bonny', 'Degema'], 2044),
    ('Kogi', ['Lokoja', 'Ajaokuta'], 2044),
    ('Borno', ['Maiduguri', 'Jere'], 2046),
]

# Ethnic column names (for reference)
ETHNIC_COLS_NAMES = [
    'Hausa', 'Fulani', 'Hausa Fulani Undiff', 'Yoruba', 'Igbo', 'Ijaw',
    'Kanuri', 'Shuwa Arab', 'Tiv', 'Ibibio', 'Efik', 'Nupe', 'Edo Bini',
    'Urhobo', 'Itsekiri', 'Isoko', 'Igala', 'Idoma', 'Berom', 'Angas',
    'Jukun', 'Ebira', 'Gwari Gbagyi', 'Ogoni', 'Ekoi', 'Tarok Yergam',
    'Bachama', 'Mumuye', 'Kuteb', 'Chamba', 'Kataf Atyap', 'Bajju',
    'Ham Jaba', 'Marghi', 'Bura', 'Kilba Huba', 'Annang', 'Oron',
    'Bolewa', 'Obolo', 'Koro', 'Eket', 'Karekare', 'Bassa', 'Gwandara',
    'Yandang', 'Kamwe', 'Gade', 'Ngizim', 'Zabarmawa', 'Bariba', 'Bata',
    'Ikwerre', 'Ngamo', 'Tuareg', 'Bade', 'Ichen', 'Yungur', 'Ganagana',
    'Tangale', 'Agatu', 'Kamuku', 'Verre', 'Mboi', 'Lala', 'Gude', 'Waja',
    'Tera', 'Kambari', 'Etche', 'Kona', 'Egun/Ogu', 'Koma', 'Pere',
    'Lunguda', 'Fali', 'Pero', 'Kare-Kare', 'Duwai', 'Ndola', 'Samba',
    'Wurkun', 'Igede', 'Pada', 'Naijin', 'Other'
]


# ============================================================================
# SECTION 2: TEXT PARSING FUNCTIONS
# ============================================================================

def safe_float(val, default=0.0):
    """Convert value to float, returning default if not possible."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace('%', '').replace(',', '')
        return float(s)
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    """Convert value to int, returning default if not possible."""
    return int(safe_float(val, default))


def parse_road_quality(text):
    """Parse road quality text to 1-10 index."""
    if not text or not isinstance(text, str):
        return 4.0  # default moderate
    t = text.lower()

    score = 5.0  # baseline

    # Negative keywords (strongest first)
    if any(w in t for w in ['no maintained', 'boat access only', 'inaccessible', 'very poor']):
        score = 1.5
    elif any(w in t for w in ['poor', 'unpaved', 'largely unpaved', 'limited paved', 'seasonal']):
        score = 3.0
    elif any(w in t for w in ['fair', 'moderate', 'partially paved', 'mixed']):
        score = 5.0
    elif any(w in t for w in ['good', 'paved roads', 'well-maintained', 'federal highway']):
        score = 7.0
    elif any(w in t for w in ['excellent', 'expressway', 'best in state', 'major highway junction']):
        score = 9.0

    # Highway presence boosts
    if re.search(r'\ba[12]\b', t):
        score = max(score, 5.5)
        score += 0.5
    if 'dual carriageway' in t:
        score += 0.5

    # Penalties
    if any(w in t for w in ['conflict', 'insecure', 'damaged', 'mined', 'bandit']):
        score -= 1.5
    if 'poorly maintained' in t:
        score -= 0.5

    # Bonuses
    if 'airport' in t:
        score += 0.5
    if 'railway' in t or 'rail' in t:
        score += 0.3
    if 'flyover' in t:
        score += 0.3

    return max(1.0, min(10.0, round(score, 1)))


def parse_market_access(text):
    """Parse market access text to 1-10 index."""
    if not text or not isinstance(text, str):
        return 4.0
    t = text.lower()

    score = 5.0

    if any(w in t for w in ['no functional', 'collapsed', 'severely limited', 'boat-dependent']):
        score = 1.5
    elif any(w in t for w in ['very limited', 'very poor', 'limited', 'remote', 'periodic']):
        score = 3.5
    elif any(w in t for w in ['moderate', 'local market', 'weekly market']):
        score = 5.0
    elif any(w in t for w in ['good', 'regional', 'commercial', 'trade hub', 'strong']):
        score = 7.0
    elif any(w in t for w in ['excellent', 'international', 'major port', 'ariaria']):
        score = 9.0

    # Distance parsing: "X km to city"
    dist_match = re.search(r'(\d+)\s*km', t)
    if dist_match:
        dist = int(dist_match.group(1))
        if dist < 20:
            score = max(score, 6.5)
        elif dist < 50:
            score = max(score, 5.0)
        elif dist > 100:
            score = min(score, 4.0)

    # State capital proximity
    if 'state capital' in t or 'capital city' in t:
        score = max(score, 7.0)

    # Penalties
    if any(w in t for w in ['disrupted', 'conflict', 'insecurity', 'banditry']):
        score -= 1.5

    return max(1.0, min(10.0, round(score, 1)))


def parse_trad_authority_influence(text):
    """Extract influence level from text. Returns (index 0-5, notes text)."""
    if not text or not isinstance(text, str):
        return 2, str(text) if text else ''
    t = text.strip()
    t_lower = t.lower()

    # Try to extract the categorical level at the start
    level = 2  # default medium
    if t_lower.startswith('dominant') or '-- dominant' in t_lower:
        level = 5
    elif t_lower.startswith('very high') or '-- very high' in t_lower:
        level = 4
    elif t_lower.startswith('high') or '-- high' in t_lower:
        level = 3
    elif t_lower.startswith('medium') or '-- medium' in t_lower or t_lower.startswith('moderate'):
        level = 2
    elif t_lower.startswith('low') or '-- low' in t_lower:
        level = 1
    elif t_lower.startswith('none') or '-- none' in t_lower or t_lower.startswith('minimal'):
        level = 0

    # Extract notes (everything after the --)
    notes = t
    dash_idx = t.find('--')
    if dash_idx >= 0 and dash_idx < 20:
        notes = t[dash_idx + 2:].strip()

    return level, notes


def parse_land_tenure(text):
    """Parse land tenure text to formalization percentage."""
    if not text or not isinstance(text, str):
        return 30.0  # default
    t = text.lower()

    if 'fcf' in t or 'federal leasehold' in t:
        return 90.0
    elif 'state leasehold' in t and 'mixed' not in t:
        return 85.0
    elif 'certificate of occupancy' in t:
        return 82.0
    elif 'mixed' in t:
        if 'state leasehold' in t:
            return 50.0
        elif 'institutional' in t:
            return 45.0
        else:
            return 40.0
    elif 'emirate customary' in t:
        return 15.0
    elif 'communal' in t or 'community' in t:
        return 15.0
    elif 'customary' in t:
        return 12.0
    else:
        return 30.0


def parse_almajiri(text, state, pct_muslim):
    """Parse and fix almajiri prevalence. Returns 0-3 index."""
    if text and isinstance(text, str):
        t = text.strip().upper()
        if t == 'HIGH':
            return 3
        elif t == 'MEDIUM':
            return 2
        elif t == 'LOW':
            return 1
        elif t == 'NONE':
            return 0

    # If text is None, numeric, or contains religious description text (data quality fix)
    if state in NORTHERN_STATES:
        if pct_muslim > 70:
            return 3  # HIGH
        elif pct_muslim > 40:
            return 2  # MEDIUM
        elif pct_muslim > 15:
            return 1  # LOW
        else:
            return 0
    else:
        return 0  # NONE for south


def parse_urban_pct(text):
    """Parse '80% urban / 20% rural' to 80.0."""
    if not text or not isinstance(text, str):
        return 20.0
    m = re.search(r'(\d+)\s*%%?\s*urban', str(text).lower())
    if m:
        return float(m.group(1))
    return 20.0


def parse_yn(text):
    """Parse Y/N to 1/0."""
    if not text:
        return 0
    return 1 if str(text).strip().upper().startswith('Y') else 0


def parse_religious_subtypes(text, pct_muslim, state):
    """Parse Religious Subtype Notes into numeric flags.
    Returns dict with keys: tijaniyya, qadiriyya, pentecostal, traditionalist_practice."""
    result = {'tijaniyya': 0, 'qadiriyya': 0, 'pentecostal': 0, 'traditionalist_practice': 0}

    if not text or not isinstance(text, str):
        # Infer from region
        if state in NORTHERN_STATES and pct_muslim > 50:
            result['tijaniyya'] = 2
            result['qadiriyya'] = 1
        elif state in SOUTHERN_STATES:
            result['pentecostal'] = 2
        return result

    t = text.lower()

    # Check if this is actually a misplaced numeric value
    try:
        float(text.strip())
        # It's a number, not real notes - use regional defaults
        if state in NORTHERN_STATES and pct_muslim > 50:
            result['tijaniyya'] = 2
            result['qadiriyya'] = 1
        elif state in SOUTHERN_STATES:
            result['pentecostal'] = 2
        return result
    except ValueError:
        pass

    # Tijaniyya
    if 'tijaniyya' in t:
        if any(w in t for w in ['dominant', 'majority', 'predominantly']):
            result['tijaniyya'] = 3
        elif any(w in t for w in ['significant', 'strong']):
            result['tijaniyya'] = 2
        else:
            result['tijaniyya'] = 1

    # Qadiriyya
    if 'qadiriyya' in t:
        if any(w in t for w in ['dominant', 'majority']):
            result['qadiriyya'] = 3
        elif any(w in t for w in ['significant', 'strong']):
            result['qadiriyya'] = 2
        else:
            result['qadiriyya'] = 1

    # If Muslim but no specific order mentioned, assign generic Sunni
    if pct_muslim > 50 and result['tijaniyya'] == 0 and result['qadiriyya'] == 0:
        if 'sunni' in t or 'maliki' in t:
            result['tijaniyya'] = 1
            result['qadiriyya'] = 1

    # Pentecostal
    if 'pentecostal' in t:
        if any(w in t for w in ['dominant', 'majority', 'strongly']):
            result['pentecostal'] = 3
        elif any(w in t for w in ['growing', 'significant', 'strong']):
            result['pentecostal'] = 2
        else:
            result['pentecostal'] = 1
    elif any(w in t for w in ['anglican', 'catholic', 'methodist', 'presbyterian']):
        result['pentecostal'] = 1  # at least some Christian presence

    # Traditionalist practice
    if any(w in t for w in ['traditionalist', 'shrine', 'oracle', 'juju', 'animist']):
        if 'strong' in t or 'persistent' in t:
            result['traditionalist_practice'] = 3
        elif 'declining' in t or 'minor' in t:
            result['traditionalist_practice'] = 1
        else:
            result['traditionalist_practice'] = 2

    return result


def parse_livelihood(text, oil_producing, urban_pct, state):
    """Parse Dominant Livelihood text into 5 percentage categories.
    Returns dict with keys: agriculture, manufacturing, extraction, services, informal.
    Must sum to 100."""
    base = {'agriculture': 50, 'manufacturing': 5, 'extraction': 0, 'services': 10, 'informal': 35}

    if not text or not isinstance(text, str):
        if urban_pct > 70:
            base = {'agriculture': 10, 'manufacturing': 15, 'extraction': 0, 'services': 30, 'informal': 45}
        if oil_producing:
            base['extraction'] = 15
            base['agriculture'] -= 10
            base['informal'] -= 5
        total = sum(base.values())
        return {k: round(v * 100 / total) for k, v in base.items()}

    t = text.lower()

    agri_kw = ['farming', 'agriculture', 'subsistence', 'pastoral', 'livestock', 'fishing',
               'crop', 'rice', 'cassava', 'yam', 'cocoa', 'palm', 'maize', 'millet', 'sorghum',
               'cattle', 'herding', 'grazing']
    mfg_kw = ['manufacturing', 'industrial', 'factory', 'artisan', 'processing', 'smelter',
              'tannery', 'textile', 'steel']
    extract_kw = ['mining', 'oil', 'gas', 'extraction', 'quarry', 'tin', 'columbite',
                  'gold', 'coal', 'limestone', 'bitumen']
    service_kw = ['commerce', 'trade', 'civil service', 'education', 'banking', 'transport',
                  'tourism', 'hospitality', 'service']
    informal_kw = ['petty trade', 'informal', 'artisan trade', 'small trading', 'hawking',
                   'motorcycle', 'okada']

    counts = {
        'agriculture': sum(1 for kw in agri_kw if kw in t),
        'manufacturing': sum(1 for kw in mfg_kw if kw in t),
        'extraction': sum(1 for kw in extract_kw if kw in t),
        'services': sum(1 for kw in service_kw if kw in t),
        'informal': sum(1 for kw in informal_kw if kw in t),
    }

    # Build base from counts
    total_kw = max(sum(counts.values()), 1)
    result = {}
    for k, c in counts.items():
        result[k] = max(5, int(c / total_kw * 80))

    # Urban adjustment
    if urban_pct > 60:
        result['agriculture'] = max(5, result['agriculture'] - 15)
        result['services'] += 10
        result['informal'] = max(10, result.get('informal', 20))

    # Oil
    if oil_producing:
        result['extraction'] = max(result['extraction'], 15)

    # Ensure minimum informal
    if result['informal'] < 15 and urban_pct < 50:
        result['informal'] = max(result['informal'], 20)

    # Normalize to 100
    total = sum(result.values())
    if total == 0:
        total = 1
    result = {k: max(1, round(v * 100 / total)) for k, v in result.items()}

    # Fix rounding
    diff = 100 - sum(result.values())
    result['informal'] += diff

    return result


def parse_resource_extraction(text, state):
    """Parse Resource Extraction Notes into structured fields.
    Returns dict: oil_active, cobalt_active, other_mining, refinery_present, intensity."""
    result = {'oil_active': 0, 'cobalt_active': 0, 'other_mining': 0,
              'refinery_present': 0, 'intensity': 0}

    if not text or not isinstance(text, str):
        return result

    t = text.lower()

    if any(w in t for w in ['oil', 'crude', 'petroleum', 'gas field', 'pipeline']):
        result['oil_active'] = 1
        result['intensity'] = 3
    if any(w in t for w in ['spill', 'producing field', 'flow station']):
        result['intensity'] = max(result['intensity'], 4)
    if any(w in t for w in ['cobalt']):
        result['cobalt_active'] = 1
    if any(w in t for w in ['mining', 'quarry', 'tin', 'gold', 'coal', 'iron',
                             'limestone', 'columbite', 'bitumen', 'lead', 'zinc',
                             'gemstone', 'granite', 'barite', 'gypite']):
        result['other_mining'] = 1
        result['intensity'] = max(result['intensity'], 2)
    if 'refinery' in t or 'smelter' in t:
        result['refinery_present'] = 1
        result['intensity'] = max(result['intensity'], 4)

    return result


# ============================================================================
# SECTION 3: DATA LOADING
# ============================================================================

def load_data(filepath):
    """Load the Excel file and return structured LGA data."""
    print(f"Loading {filepath}...")
    wb = load_workbook(filepath, data_only=True)
    ws = wb['LGA_DATA']

    # Read header rows
    row1 = [cell.value for cell in ws[1]]  # group headers
    row2 = [cell.value for cell in ws[2]]  # column names

    # Read all LGA data rows
    lgas = []
    for row_idx in range(3, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=c + 1).value for c in range(ws.max_column)]
        if row_data[0] is None and row_data[1] is None:
            continue
        lgas.append(row_data)

    print(f"  Loaded {len(lgas)} LGAs with {len(row2)} columns")

    # Read metadata
    ws_meta = wb['METADATA']
    metadata = []
    for row in ws_meta.iter_rows(min_row=1, values_only=True):
        metadata.append(list(row))

    wb.close()
    return lgas, row2, row1, metadata


# ============================================================================
# SECTION 4: LGA CLASS - structured data per LGA
# ============================================================================

class LGA:
    """Represents a single LGA with all its data fields."""

    def __init__(self, row_data, col_names):
        self.raw = list(row_data)  # preserve original

        # Identity
        self.state = str(row_data[COL['state']] or '').strip()
        self.lga_name = str(row_data[COL['lga_name']] or '').strip()
        self.colonial_region = str(row_data[COL['colonial_region']] or '').strip()
        self.terrain = str(row_data[COL['terrain']] or '').strip()

        # Demographics
        self.population = safe_float(row_data[COL['population']])
        self.pop_density = safe_float(row_data[COL['pop_density']])
        self.median_age = safe_float(row_data[COL['median_age']], 21.0)
        self.pct_under_30 = safe_float(row_data[COL['pct_under_30']], 65.0)

        # Ethnicity (86 columns)
        self.ethnic = []
        for i in range(COL['ethnic_start'], COL['ethnic_end'] + 1):
            self.ethnic.append(safe_float(row_data[i]))
        # Indices within ethnic array: Pada=83, Naijin=84, Other=85
        self.pada_idx = 83
        self.naijin_idx = 84
        self.other_idx = 85

        # Religion
        self.pct_muslim = safe_float(row_data[COL['pct_muslim']])
        self.pct_christian = safe_float(row_data[COL['pct_christian']])
        self.pct_traditionalist = safe_float(row_data[COL['pct_traditionalist']])

        # Religious subtypes (raw text preserved, will be parsed)
        self.religious_notes_raw = row_data[COL['religious_subtype_notes']]
        self.almajiri_raw = row_data[COL['almajiri_prevalence']]

        # Urbanization
        self.major_urban_center = row_data[COL['major_urban_center']]
        self.urban_pct = parse_urban_pct(row_data[COL['urban_rural_split']])

        # Economic
        self.oil_producing = parse_yn(row_data[COL['oil_producing']])
        self.livelihood_raw = row_data[COL['dominant_livelihood']]
        self.resource_raw = row_data[COL['resource_extraction_notes']]
        self.poverty_rate = safe_float(row_data[COL['poverty_rate']], 50.0)
        self.unemployment = safe_float(row_data[COL['unemployment_rate']], 20.0)
        self.youth_unemployment = safe_float(row_data[COL['youth_unemployment']], 30.0)

        # Infrastructure
        self.access_electricity = safe_float(row_data[COL['access_electricity']], 30.0)
        self.access_water = safe_float(row_data[COL['access_water']], 40.0)
        self.access_healthcare = safe_float(row_data[COL['access_healthcare']], 30.0)
        self.road_quality_raw = row_data[COL['road_quality']]
        self.market_access_raw = row_data[COL['market_access']]

        # Education
        self.adult_literacy = safe_float(row_data[COL['adult_literacy']], 50.0)
        self.male_literacy = safe_float(row_data[COL['male_literacy']], 55.0)
        self.female_literacy = safe_float(row_data[COL['female_literacy']], 45.0)
        self.primary_enrollment = safe_float(row_data[COL['primary_enrollment']], 60.0)
        self.secondary_enrollment = safe_float(row_data[COL['secondary_enrollment']], 40.0)
        self.gender_parity = safe_float(row_data[COL['gender_parity']], 0.85)
        self.out_of_school = safe_float(row_data[COL['out_of_school']], 30.0)
        self.num_secondary_schools = safe_int(row_data[COL['num_secondary_schools']], 5)
        self.tertiary_institution = parse_yn(row_data[COL['tertiary_institution']])

        # Political
        self.trad_authority = row_data[COL['trad_authority']]
        self.trad_influence_raw = row_data[COL['trad_authority_influence']]
        self.land_tenure_raw = row_data[COL['predominant_land_tenure']]

        # Connectivity
        self.mobile_penetration = safe_float(row_data[COL['mobile_penetration']], 40.0)
        self.internet_access = safe_float(row_data[COL['internet_access']], 15.0)
        self.data_notes = row_data[COL['data_notes']]

        # ---- Parse text columns into numeric ----
        self.road_quality_idx = parse_road_quality(self.road_quality_raw)
        self.market_access_idx = parse_market_access(self.market_access_raw)
        ti_level, ti_notes = parse_trad_authority_influence(self.trad_influence_raw)
        self.trad_authority_idx = ti_level
        self.trad_authority_notes = ti_notes
        self.land_formalization = parse_land_tenure(self.land_tenure_raw)
        self.land_tenure_notes = str(self.land_tenure_raw) if self.land_tenure_raw else ''
        self.almajiri_idx = parse_almajiri(self.almajiri_raw, self.state, self.pct_muslim)

        # Religious subtypes
        rs = parse_religious_subtypes(self.religious_notes_raw, self.pct_muslim, self.state)
        self.tijaniyya = rs['tijaniyya']
        self.qadiriyya = rs['qadiriyya']
        self.pentecostal = rs['pentecostal']
        self.trad_practice = rs['traditionalist_practice']
        self.al_shahid_influence = 0.0
        self.religious_notes = str(self.religious_notes_raw) if self.religious_notes_raw else ''

        # Livelihood decomposition
        liv = parse_livelihood(self.livelihood_raw, self.oil_producing,
                               self.urban_pct, self.state)
        self.liv_agriculture = liv['agriculture']
        self.liv_manufacturing = liv['manufacturing']
        self.liv_extraction = liv['extraction']
        self.liv_services = liv['services']
        self.liv_informal = liv['informal']
        self.livelihood_notes = str(self.livelihood_raw) if self.livelihood_raw else ''

        # Resource extraction decomposition
        res = parse_resource_extraction(self.resource_raw, self.state)
        self.oil_extraction_active = res['oil_active']
        self.cobalt_extraction_active = res['cobalt_active']
        self.other_mining_active = res['other_mining']
        self.refinery_present = res['refinery_present']
        self.extraction_intensity = res['intensity']

        # ---- NEW COLUMNS (initialized) ----
        az = AZ_MAP.get(self.state, (0, 'Unassigned'))
        self.admin_zone = az[0]
        self.az_name = az[1]
        self.conflict_history = 'SPARED'
        self.federal_control = 'FULL'
        self.bic_effectiveness = 0.0

        # Economic new
        self.gdp_per_capita = 800.0
        self.gini_proxy = 0.35
        self.chinese_economic_presence = 0.0
        self.rail_corridor = 0
        self.planned_city = 0
        self.refinery_zone = 0
        self.housing_affordability = 7.0

        # Demographic new
        self.fertility_rate = 5.5 if self.state in NORTHERN_STATES else 4.5
        self.bio_enhancement_pct = 0.0

        # Cultural new
        self.english_prestige = 5.0 if self.state in SOUTHERN_STATES else 3.0
        self.mandarin_presence = 0.0
        self.arabic_prestige = 3.0 if (self.state in NORTHERN_STATES and self.pct_muslim > 50) else 0.5

        # Internal simulation state (not output directly)
        self._conflict_status = 'STABLE'  # STABLE/LOW_CONFLICT/HIGH_CONFLICT/OCCUPIED/AL_SHAHID_CONTROL
        self._development_effectiveness = 1.0  # multiplier 0-1

        # Flags
        self.is_northern = self.state in NORTHERN_STATES
        self.is_southern = self.state in SOUTHERN_STATES
        self.is_lagos = self.state == 'Lagos'
        self.is_fct = self.state == 'FCT'

    def conflict_effectiveness(self):
        """Return development program effectiveness based on conflict status."""
        eff = {
            'STABLE': 1.0, 'LOW_CONFLICT': 0.8, 'HIGH_CONFLICT': 0.4,
            'OCCUPIED': 0.0, 'AL_SHAHID_CONTROL': 0.0
        }
        return eff.get(self._conflict_status, 0.5)

    def normalize_ethnic(self):
        """Ensure ethnic percentages sum to ~100."""
        total = sum(self.ethnic)
        if total > 0 and abs(total - 100) > 0.5:
            factor = 100.0 / total
            self.ethnic = [e * factor for e in self.ethnic]

    def normalize_religion(self):
        """Ensure religious percentages sum to ~100."""
        total = self.pct_muslim + self.pct_christian + self.pct_traditionalist
        if total > 0 and abs(total - 100) > 0.5:
            factor = 100.0 / total
            self.pct_muslim *= factor
            self.pct_christian *= factor
            self.pct_traditionalist *= factor

    def normalize_livelihood(self):
        """Ensure livelihood percentages sum to 100."""
        total = (self.liv_agriculture + self.liv_manufacturing +
                 self.liv_extraction + self.liv_services + self.liv_informal)
        if total > 0 and abs(total - 100) > 0.5:
            factor = 100.0 / total
            self.liv_agriculture *= factor
            self.liv_manufacturing *= factor
            self.liv_extraction *= factor
            self.liv_services *= factor
            self.liv_informal *= factor

    def clamp_percentages(self):
        """Ensure all percentage fields are in valid ranges."""
        pct_fields = [
            'pct_muslim', 'pct_christian', 'pct_traditionalist',
            'poverty_rate', 'unemployment', 'youth_unemployment',
            'access_electricity', 'access_water', 'access_healthcare',
            'adult_literacy', 'male_literacy', 'female_literacy',
            'primary_enrollment', 'secondary_enrollment',
            'out_of_school', 'mobile_penetration', 'internet_access',
            'urban_pct', 'bio_enhancement_pct', 'land_formalization',
        ]
        for field in pct_fields:
            val = getattr(self, field)
            setattr(self, field, max(0.0, min(100.0, val)))

        # Ethnic
        for i in range(len(self.ethnic)):
            self.ethnic[i] = max(0.0, min(100.0, self.ethnic[i]))

        # Livelihood
        for field in ['liv_agriculture', 'liv_manufacturing', 'liv_extraction',
                       'liv_services', 'liv_informal']:
            setattr(self, field, max(0.0, min(100.0, getattr(self, field))))

        # Indices
        self.road_quality_idx = max(1.0, min(10.0, self.road_quality_idx))
        self.market_access_idx = max(1.0, min(10.0, self.market_access_idx))
        self.trad_authority_idx = max(0, min(5, self.trad_authority_idx))
        self.almajiri_idx = max(0, min(5, self.almajiri_idx))
        self.al_shahid_influence = max(0.0, min(5.0, self.al_shahid_influence))
        self.bic_effectiveness = max(0.0, min(10.0, self.bic_effectiveness))
        self.chinese_economic_presence = max(0.0, min(10.0, self.chinese_economic_presence))
        self.housing_affordability = max(0.0, min(10.0, self.housing_affordability))
        self.gini_proxy = max(0.0, min(1.0, self.gini_proxy))
        self.fertility_rate = max(1.0, min(5.0, self.fertility_rate))
        self.gender_parity = max(0.0, min(1.5, self.gender_parity))


# ============================================================================
# SECTION 5: ECONOMIC MODEL
# ============================================================================

def interpolate_gdp(year):
    """Interpolate national GDP per capita for a given year using log-linear interpolation."""
    years = sorted(GDP_TRAJECTORY.keys())
    if year <= years[0]:
        return GDP_TRAJECTORY[years[0]]
    if year >= years[-1]:
        return GDP_TRAJECTORY[years[-1]]

    for i in range(len(years) - 1):
        if years[i] <= year <= years[i + 1]:
            y0, y1 = years[i], years[i + 1]
            g0, g1 = GDP_TRAJECTORY[y0], GDP_TRAJECTORY[y1]
            # Log-linear interpolation for exponential growth
            if g0 > 0 and g1 > 0:
                log_g0 = math.log(g0)
                log_g1 = math.log(g1)
                frac = (year - y0) / (y1 - y0)
                return math.exp(log_g0 + frac * (log_g1 - log_g0))
            else:
                frac = (year - y0) / (y1 - y0)
                return g0 + frac * (g1 - g0)
    return 800


def compute_lga_gdp(lga, national_gdp):
    """Compute LGA-level GDP per capita based on national average and local factors."""
    multiplier = 1.0

    # Urban premium
    if lga.urban_pct > 80:
        multiplier *= 1.6
    elif lga.urban_pct > 50:
        multiplier *= 1.2
    elif lga.urban_pct < 20:
        multiplier *= 0.5

    # Lagos premium
    if lga.is_lagos:
        multiplier *= 1.8

    # Conflict penalty
    status_mult = {
        'STABLE': 1.0, 'LOW_CONFLICT': 0.85, 'HIGH_CONFLICT': 0.5,
        'OCCUPIED': 0.2, 'AL_SHAHID_CONTROL': 0.3
    }
    multiplier *= status_mult.get(lga._conflict_status, 0.7)

    # Sector mix: extraction and services pay more
    sector_bonus = (lga.liv_extraction * 0.02 + lga.liv_services * 0.01 +
                    lga.liv_manufacturing * 0.008 - lga.liv_agriculture * 0.005)
    multiplier *= (1.0 + sector_bonus / 100)

    # North-South gap
    if lga.is_northern:
        multiplier *= 0.7
    elif lga.is_southern:
        multiplier *= 1.15

    # Planned city / refinery / rail
    if lga.planned_city:
        multiplier *= 1.3
    if lga.refinery_zone:
        multiplier *= 1.2
    if lga.rail_corridor == 2:
        multiplier *= 1.1
    elif lga.rail_corridor == 1:
        multiplier *= 1.05

    # Oil producing
    if lga.oil_producing:
        multiplier *= 1.1

    return national_gdp * multiplier


def compute_gini(lga):
    """Estimate local Gini coefficient."""
    base = 0.35

    # Urban areas have higher inequality
    if lga.urban_pct > 80:
        base += 0.1
    elif lga.urban_pct > 50:
        base += 0.05

    # Lagos housing crisis
    if lga.is_lagos:
        base += 0.08

    # Planned cities: inequality from pada/local divide
    if lga.planned_city:
        base += 0.05

    # Homogeneous rural areas: lower inequality
    if lga.urban_pct < 25 and lga.poverty_rate > 60:
        base -= 0.05  # everyone equally poor

    # Al-Shahid zones: low inequality (everyone poor)
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        base = 0.25

    return max(0.15, min(0.65, base))


def compute_housing_affordability(lga, year):
    """Compute housing affordability (10=very affordable, 0=crisis)."""
    base = 7.0

    # Rural areas: affordable
    if lga.urban_pct < 25:
        base = 8.5
    elif lga.urban_pct > 80:
        base = 5.0

    # Lagos crisis escalates over time
    if lga.is_lagos:
        decline = max(0, (year - 2035)) * 0.15
        base = max(1.0, 5.0 - decline)

    # Planned cities: initially affordable then pressure builds
    if lga.planned_city:
        years_since = max(0, year - 2040)
        base = max(2.0, 6.0 - years_since * 0.2)

    # High pop density
    if lga.pop_density > 2000:
        base -= 1.0
    elif lga.pop_density > 500:
        base -= 0.3

    return max(0.0, min(10.0, base))


# ============================================================================
# SECTION 6: CONFLICT STATUS MODEL
# ============================================================================

# LGAs in specific conflict categories
ISWAP_STATES = {'Borno', 'Yobe', 'Adamawa'}
BANDIT_STATES = {'Zamfara', 'Katsina', 'Kaduna', 'Niger', 'Sokoto', 'Kebbi'}
ASS_INVASION_STATES = {'Sokoto', 'Zamfara', 'Kebbi', 'Niger'}  # plus FCT

def update_conflict_status(lga, year, lgas_dict=None):
    """Update conflict status for an LGA based on the year and timeline events."""
    state = lga.state
    name = lga.lga_name.lower()

    # --- 2026-2027: Baseline ---
    if year <= 2027:
        if state in ISWAP_STATES:
            if state == 'Borno':
                lga._conflict_status = 'HIGH_CONFLICT'
            elif state == 'Yobe':
                lga._conflict_status = 'LOW_CONFLICT'
            elif state == 'Adamawa':
                lga._conflict_status = 'LOW_CONFLICT'
        elif state in BANDIT_STATES:
            lga._conflict_status = 'LOW_CONFLICT'
        return

    # --- 2028-2029: Escalation ---
    if year in (2028, 2029):
        if state in ISWAP_STATES:
            lga._conflict_status = 'HIGH_CONFLICT'
        elif state in BANDIT_STATES:
            if lga._conflict_status == 'STABLE':
                lga._conflict_status = 'LOW_CONFLICT'
            elif lga._conflict_status == 'LOW_CONFLICT':
                lga._conflict_status = 'HIGH_CONFLICT'
        return

    # --- 2030: ASS Invasion ---
    if year == 2030:
        if state == 'Zamfara':
            lga._conflict_status = 'OCCUPIED'
        elif state in ('Sokoto',):
            # Sokoto spared (non-aggression pact) but shaken
            lga._conflict_status = 'LOW_CONFLICT'
        elif state in ('Kebbi',):
            lga._conflict_status = 'HIGH_CONFLICT'
        elif state == 'Niger':
            # Corridor toward Abuja
            lga._conflict_status = 'HIGH_CONFLICT'
        elif state == 'FCT':
            lga._conflict_status = 'HIGH_CONFLICT'
        elif state in ISWAP_STATES:
            lga._conflict_status = 'HIGH_CONFLICT'
        return

    # --- 2031: Liberation ---
    if year == 2031:
        if state == 'Zamfara':
            lga._conflict_status = 'HIGH_CONFLICT'  # liberated but damaged
        elif state == 'FCT':
            lga._conflict_status = 'LOW_CONFLICT'  # recaptured
        elif state in ('Kebbi', 'Niger'):
            lga._conflict_status = 'HIGH_CONFLICT'  # still recovering
        elif state in ISWAP_STATES:
            lga._conflict_status = 'HIGH_CONFLICT'
        elif state == 'Sokoto':
            lga._conflict_status = 'LOW_CONFLICT'
        return

    # --- 2032-2039: Post-war recovery (gradual) ---
    if 2032 <= year <= 2039:
        if lga._conflict_status == 'OCCUPIED':
            lga._conflict_status = 'HIGH_CONFLICT'
        elif lga._conflict_status == 'HIGH_CONFLICT':
            # Gradual recovery
            recovery_years = year - 2031
            if state in ('FCT', 'Lagos'):
                lga._conflict_status = 'STABLE'
            elif recovery_years >= 5 and state not in ISWAP_STATES:
                lga._conflict_status = 'LOW_CONFLICT'
            elif recovery_years >= 8 and state not in ISWAP_STATES:
                lga._conflict_status = 'STABLE'
        elif lga._conflict_status == 'LOW_CONFLICT':
            if year >= 2034 and state not in ISWAP_STATES:
                lga._conflict_status = 'STABLE'

        # ISWAP states recover slower
        if state in ISWAP_STATES and year < 2038:
            if lga._conflict_status == 'STABLE':
                lga._conflict_status = 'LOW_CONFLICT'
        return

    # --- 2040: Kano falls to al-Shahid ---
    if year == 2040:
        if state == 'Kano':
            lga._conflict_status = 'AL_SHAHID_CONTROL'
        elif state in ('Jigawa', 'Bauchi', 'Kaduna'):
            # Neighboring states affected
            if lga._conflict_status == 'STABLE':
                lga._conflict_status = 'LOW_CONFLICT'
        return

    # --- 2041-2051: Al-Shahid era ---
    if 2041 <= year <= 2051:
        if state == 'Kano':
            lga._conflict_status = 'AL_SHAHID_CONTROL'
        elif state in ISWAP_STATES:
            if lga._conflict_status in ('STABLE', 'LOW_CONFLICT'):
                lga._conflict_status = 'LOW_CONFLICT'
        # A2 highway attacks (2043)
        if year == 2043 and state in ('Kaduna', 'Niger', 'FCT'):
            if lga._conflict_status == 'STABLE':
                lga._conflict_status = 'LOW_CONFLICT'

        # Most of south stays stable
        if lga.is_southern and lga._conflict_status in ('LOW_CONFLICT', 'HIGH_CONFLICT'):
            if year >= 2042:
                lga._conflict_status = 'STABLE'
        return

    # --- 2052+: Peace talks, gradual recovery ---
    if year >= 2052:
        if state == 'Kano':
            if year <= 2054:
                lga._conflict_status = 'HIGH_CONFLICT'
            elif year <= 2056:
                lga._conflict_status = 'LOW_CONFLICT'
            else:
                lga._conflict_status = 'LOW_CONFLICT'  # nominal by 2058
        elif state in ISWAP_STATES:
            if year >= 2055 and state != 'Borno':
                lga._conflict_status = 'STABLE'
            elif state == 'Borno':
                lga._conflict_status = 'LOW_CONFLICT'
        else:
            if lga._conflict_status != 'STABLE':
                lga._conflict_status = 'STABLE'


def update_federal_control(lga, year):
    """Update Federal Control categorical based on conflict status."""
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.federal_control = 'NONE'
    elif lga._conflict_status == 'OCCUPIED':
        lga.federal_control = 'NONE'
    elif lga._conflict_status == 'HIGH_CONFLICT':
        lga.federal_control = 'PARTIAL'
    elif lga._conflict_status == 'LOW_CONFLICT':
        if year >= 2052 and lga.state == 'Kano':
            lga.federal_control = 'NOMINAL'
        else:
            lga.federal_control = 'PARTIAL'
    else:
        lga.federal_control = 'FULL'

    # Sokoto exception: traditional authority
    if lga.state == 'Sokoto' and year >= 2030 and year <= 2040:
        if lga.federal_control == 'FULL':
            lga.federal_control = 'PARTIAL'

    # 2058 final
    if year == 2058:
        if lga.state == 'Kano':
            lga.federal_control = 'NOMINAL'
        elif lga.state == 'Borno' and lga._conflict_status == 'LOW_CONFLICT':
            lga.federal_control = 'PARTIAL'


def update_conflict_history(lga, year):
    """Update the Conflict History categorical based on worst status experienced."""
    mapping = {
        'STABLE': 'SPARED',
        'LOW_CONFLICT': 'LOW_IMPACT',
        'HIGH_CONFLICT': 'MODERATE',
        'OCCUPIED': 'OCCUPATION',
        'AL_SHAHID_CONTROL': 'AL_SHAHID_CONTROL',
    }
    current = mapping.get(lga._conflict_status, 'LOW_IMPACT')

    severity = {'SPARED': 0, 'LOW_IMPACT': 1, 'MODERATE': 2, 'SEVERE': 3,
                'OCCUPATION': 4, 'AL_SHAHID_CONTROL': 5}

    if severity.get(current, 0) > severity.get(lga.conflict_history, 0):
        lga.conflict_history = current

    # Upgrade MODERATE to SEVERE for prolonged conflict
    if lga.conflict_history == 'MODERATE' and lga._conflict_status == 'HIGH_CONFLICT':
        lga.conflict_history = 'SEVERE'


# ============================================================================
# SECTION 7: MIGRATION MODEL
# ============================================================================

def compute_migration(lga, year, all_lgas):
    """Compute net migration rate for an LGA. Returns (emigration_rate, immigration_rate).
    Rates are fractions of current population (e.g., 0.03 = 3%)."""
    emigration = 0.0
    immigration = 0.0

    # ---- Push factors ----
    if lga._conflict_status in ('HIGH_CONFLICT', 'OCCUPIED'):
        emigration += 0.04
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        # Non-Muslims flee
        non_muslim_share = (lga.pct_christian + lga.pct_traditionalist) / 100.0
        emigration += 0.08 * non_muslim_share + 0.03  # some Muslims flee too
    if lga._conflict_status == 'OCCUPIED':
        emigration += 0.06

    if lga.poverty_rate > 70 and lga.unemployment > 30:
        emigration += 0.008
    elif lga.poverty_rate > 50 and lga.unemployment > 20:
        emigration += 0.004

    # War years
    if year in (2030, 2031) and lga.state in ASS_INVASION_STATES:
        emigration += 0.07

    # ---- Pull factors ----
    if lga.urban_pct > 60 and lga.poverty_rate < 40:
        immigration += 0.015
    if lga.is_lagos:
        immigration += 0.025
    if lga.planned_city and year > 2035:
        immigration += 0.03
    if lga.refinery_zone:
        immigration += 0.015
    if lga.is_fct and year < 2032:
        immigration += 0.01  # Pre-capital-move
    if lga.rail_corridor >= 1 and year >= 2035:
        immigration += 0.005

    return emigration, immigration


def apply_migration(lgas, year):
    """Apply migration across all LGAs for one year.
    Computes net population changes and adjusts ethnic/religious composition."""
    total_emigration = {}  # state -> total leaving
    total_immigration = {}  # state -> total arriving
    lga_emigration = {}
    lga_immigration = {}

    # Phase 1: compute rates
    for lga in lgas:
        emig, immig = compute_migration(lga, year, lgas)
        pop_leaving = lga.population * emig
        pop_arriving = lga.population * immig  # base attraction
        lga_emigration[lga.lga_name] = pop_leaving
        lga_immigration[lga.lga_name] = pop_arriving

    # Phase 2: balance - total emigrating should roughly equal total immigrating
    total_emig = sum(lga_emigration.values())
    total_immig = sum(lga_immigration.values())

    if total_immig > 0:
        # Scale immigration to match emigration (conservation of population in migration)
        scale = total_emig / total_immig if total_immig > 0 else 1.0
        scale = min(scale, 2.0)  # cap
    else:
        scale = 0.0

    # Phase 3: Apply
    for lga in lgas:
        emig_pop = lga_emigration[lga.lga_name]
        immig_pop = lga_immigration[lga.lga_name] * scale

        net = immig_pop - emig_pop
        lga.population = max(1000, lga.population + net)

        # Ethnic shift from migration:
        # Immigration brings a mix roughly proportional to national urban average
        # (simplified: pada/naijin for pull-cities, existing mix otherwise)
        if immig_pop > 0 and lga.population > 0:
            immig_frac = immig_pop / lga.population
            if immig_frac > 0.01:
                # Immigrants to Lagos / planned cities bring pada/naijin
                if lga.is_lagos or lga.planned_city:
                    pada_immig = immig_frac * 0.08  # 8% of immigrants are pada
                    naijin_immig = immig_frac * 0.03
                    lga.ethnic[lga.pada_idx] += pada_immig
                    lga.ethnic[lga.naijin_idx] += naijin_immig
                    # Reduce others proportionally
                    total_other = sum(lga.ethnic) - lga.ethnic[lga.pada_idx] - lga.ethnic[lga.naijin_idx]
                    if total_other > 0:
                        reduce_factor = (100 - lga.ethnic[lga.pada_idx] - lga.ethnic[lga.naijin_idx]) / total_other
                        for i in range(len(lga.ethnic)):
                            if i not in (lga.pada_idx, lga.naijin_idx):
                                lga.ethnic[i] *= reduce_factor

        # Special: anti-Muslim pogrom in Lagos 2040
        if year == 2040 and lga.is_lagos:
            muslim_loss = min(lga.pct_muslim * 0.3, 3.0)
            lga.pct_muslim -= muslim_loss
            lga.pct_christian += muslim_loss * 0.7
            lga.pct_traditionalist += muslim_loss * 0.3

        lga.normalize_ethnic()
        lga.normalize_religion()

    # Update population density
    for lga in lgas:
        if lga.pop_density > 0 and lga.population > 0:
            # Estimate area from original density and population
            original_area = lga.raw[COL['population']] or lga.population
            original_density = lga.raw[COL['pop_density']] or 100
            if original_density > 0:
                area = safe_float(original_area) / safe_float(original_density)
                if area > 0:
                    lga.pop_density = lga.population / area


# ============================================================================
# SECTION 8: DEMOGRAPHIC UPDATE FUNCTIONS
# ============================================================================

def update_population_growth(lga, year):
    """Apply natural population growth based on fertility rate.
    Target: ~131M in 2026 -> ~280-320M in 2058 (avg ~2.7% growth early, slowing)."""
    # Convert fertility to growth rate
    # Nigeria's current growth ~2.5% with TFR ~5.2
    # TFR 5.5 -> ~3.2%, TFR 3.0 -> ~1.5%, TFR 2.0 -> ~0.5%, TFR 1.3 -> ~-0.1%
    growth_rate = (lga.fertility_rate - 1.1) * 0.0078 + 0.002
    # Longevity improvement bonus (healthcare, bio-enhancement)
    if year > 2035:
        growth_rate += 0.003 * lga.conflict_effectiveness()
    if year > 2047 and lga.bio_enhancement_pct > 10:
        growth_rate += 0.001  # bio-enhancement reduces mortality
    growth_rate = max(-0.002, min(0.035, growth_rate))
    lga.population *= (1.0 + growth_rate)
    lga.population = max(1000, lga.population)


def update_fertility(lga, year):
    """Update fertility rate using target-convergence approach.
    Targets by 2058: S-urban 1.35, S-rural 1.9, N-urban 2.3, N-rural 3.2."""
    # Determine target for this LGA
    if lga.is_southern and lga.urban_pct > 60:
        target = 1.35
    elif lga.is_southern:
        target = 1.9
    elif lga.is_northern and lga.urban_pct > 50:
        target = 2.3
    else:
        target = 3.2

    # Convergence speed depends on era and conflict
    eff = lga.conflict_effectiveness()
    years_remaining = max(1, END_YEAR - year)

    # Gap between current and target
    gap = lga.fertility_rate - target

    if gap <= 0:
        return  # already at or below target

    # Converge: close a fraction of the gap each year
    # Speed: slow 2026-2032 (state collapse), medium 2032-2040, fast 2040+
    if year < 2032:
        speed = 0.02  # close 2% of gap per year (state collapse era)
    elif year < 2040:
        speed = 0.05
    elif year < 2050:
        speed = 0.09
    else:
        speed = 0.14

    # Conflict reduces speed
    speed *= (0.3 + 0.7 * eff)

    # Southern areas converge faster
    if lga.is_southern:
        speed *= 1.3

    # Pada effect: pada communities push fertility down faster
    if lga.ethnic[lga.pada_idx] > 2:
        speed *= 1.2

    decline = gap * speed
    lga.fertility_rate -= decline

    # Hard floors
    if lga.is_southern and lga.urban_pct > 60:
        lga.fertility_rate = max(1.3, lga.fertility_rate)
    elif lga.is_southern:
        lga.fertility_rate = max(1.7, lga.fertility_rate)
    elif lga.is_northern and lga.urban_pct > 50:
        lga.fertility_rate = max(2.0, lga.fertility_rate)
    else:
        lga.fertility_rate = max(2.8, lga.fertility_rate)

    lga.fertility_rate = min(5.5, lga.fertility_rate)


def update_age_structure(lga, year):
    """Update median age and % under 30 based on fertility trends."""
    # Median age rises as fertility drops (and vice versa)
    # Target by 2058: South 32-37, North 19-24
    years_elapsed = year - 2026
    total_years = 32

    if lga.is_southern:
        if lga.urban_pct > 60:
            target_median = 36.0
            target_under30 = 40.0
        else:
            target_median = 32.0
            target_under30 = 45.0
    else:
        if lga.urban_pct > 50:
            target_median = 24.0
            target_under30 = 60.0
        else:
            target_median = 20.0
            target_under30 = 66.0

    # Move toward targets using logistic curve
    progress = years_elapsed / total_years
    # Sigmoid: slow start, fast middle, slow end
    sigmoid = 1.0 / (1.0 + math.exp(-8 * (progress - 0.5)))

    start_median = lga.raw[COL['median_age']]
    start_under30 = lga.raw[COL['pct_under_30']]
    if start_median is None:
        start_median = 21.0
    if start_under30 is None:
        start_under30 = 65.0
    start_median = safe_float(start_median, 21.0)
    start_under30 = safe_float(start_under30, 65.0)

    lga.median_age = start_median + (target_median - start_median) * sigmoid
    lga.pct_under_30 = start_under30 + (target_under30 - start_under30) * sigmoid


def update_urbanization(lga, year):
    """Urbanization increases over time."""
    eff = lga.conflict_effectiveness()

    if lga.urban_pct >= 98:
        return  # already fully urban

    base_rate = 0.5  # baseline pct points per year
    if lga.is_lagos:
        base_rate = 0.3  # already very urban
    elif lga.is_southern and lga.urban_pct > 50:
        base_rate = 0.7
    elif lga.is_northern and lga.urban_pct < 20:
        base_rate = 0.4

    # Conflict slows urbanization
    lga.urban_pct += base_rate * eff
    lga.urban_pct = min(99.0, lga.urban_pct)


def update_pada_naijin(lga, year):
    """Update Pada and Naijin percentages based on timeline events."""
    pada_idx = lga.pada_idx
    naijin_idx = lga.naijin_idx

    # --- Pada immigration ---
    if year < 2029:
        return  # No pada yet

    pada_rate = 0.0
    naijin_rate = 0.0

    if year == 2029:
        # Trickle into Lagos
        if lga.is_lagos:
            pada_rate = 0.05
    elif year <= 2032:
        if lga.is_lagos:
            pada_rate = 0.3
        elif lga.is_fct:
            pada_rate = 0.1
    elif year <= 2040:
        # Steady growth
        if lga.is_lagos:
            pada_rate = 0.4
        elif lga.planned_city:
            pada_rate = 0.3
        elif lga.is_southern and lga.urban_pct > 50:
            pada_rate = 0.1
        elif lga.is_fct:
            pada_rate = 0.15
        elif lga.is_northern and lga.urban_pct > 60:
            pada_rate = 0.03

        # Naijin start appearing 2038
        if year >= 2038:
            if lga.is_lagos:
                naijin_rate = 0.15
            elif lga.planned_city:
                naijin_rate = 0.1
            elif lga.rail_corridor >= 1 and lga.chinese_economic_presence > 3:
                naijin_rate = 0.05
    elif year <= 2050:
        if lga.is_lagos:
            pada_rate = 0.35
            naijin_rate = 0.2
        elif lga.planned_city:
            pada_rate = 0.25
            naijin_rate = 0.15
        elif lga.is_southern and lga.urban_pct > 50:
            pada_rate = 0.1
            naijin_rate = 0.05
        elif lga.is_fct:
            pada_rate = 0.15
            naijin_rate = 0.08
        elif lga.is_northern and lga.urban_pct > 60:
            pada_rate = 0.02
            naijin_rate = 0.01
    else:
        # 2050-2058: slowing slightly
        if lga.is_lagos:
            pada_rate = 0.25
            naijin_rate = 0.15
        elif lga.planned_city:
            pada_rate = 0.15
            naijin_rate = 0.1
        elif lga.is_southern and lga.urban_pct > 50:
            pada_rate = 0.08
            naijin_rate = 0.03
        elif lga.is_fct:
            pada_rate = 0.1
            naijin_rate = 0.05
        elif lga.is_northern and lga.urban_pct > 60:
            pada_rate = 0.01
            naijin_rate = 0.005

    lga.ethnic[pada_idx] += pada_rate
    lga.ethnic[naijin_idx] += naijin_rate

    # Reduce "Other" to compensate
    lga.ethnic[lga.other_idx] -= (pada_rate + naijin_rate) * 0.5
    if lga.ethnic[lga.other_idx] < 0.5:
        lga.ethnic[lga.other_idx] = 0.5

    lga.normalize_ethnic()


# ============================================================================
# SECTION 9: INFRASTRUCTURE & DEVELOPMENT UPDATES
# ============================================================================

def update_infrastructure(lga, year):
    """Update infrastructure indices based on year, programs, and conflict."""
    eff = lga.conflict_effectiveness()

    # --- Electricity ---
    if lga.is_southern:
        lga.access_electricity += 1.5 * eff
    else:
        lga.access_electricity += 0.7 * eff

    # Shiroro dam attack 2045: temporary blackout in Niger Delta
    if year == 2045 and lga.state in ('Delta', 'Rivers', 'Bayelsa'):
        lga.access_electricity -= 7.0
    elif year == 2046 and lga.state in ('Delta', 'Rivers', 'Bayelsa'):
        lga.access_electricity += 5.0  # recovery

    # --- Water ---
    if lga.is_southern:
        lga.access_water += 1.0 * eff
    else:
        lga.access_water += 0.5 * eff

    # --- Healthcare ---
    if lga.is_southern:
        lga.access_healthcare += 0.8 * eff
    else:
        lga.access_healthcare += 0.4 * eff

    # BUA pharma 2034+
    if year >= 2034 and lga.admin_zone == 8:
        lga.access_healthcare += 0.2 * eff

    # Refinery zone boost
    if lga.refinery_zone and year >= 2044:
        lga.access_electricity += 0.5
        lga.access_water += 0.3

    # --- Road Quality ---
    if year >= 2037:
        # Road improvement program
        if lga.is_southern:
            lga.road_quality_idx += 0.35 * eff
        elif lga.is_northern and lga._conflict_status == 'STABLE':
            lga.road_quality_idx += 0.2 * eff
        elif lga.is_northern:
            lga.road_quality_idx += 0.1 * eff
        # Diminishing returns after 2050
        if year >= 2050:
            pass  # already slowing via eff

    # Rail corridor boost to road quality
    if lga.rail_corridor >= 1:
        lga.road_quality_idx += 0.05

    # Conflict degrades roads
    if lga._conflict_status == 'HIGH_CONFLICT':
        lga.road_quality_idx -= 0.15
    elif lga._conflict_status in ('OCCUPIED', 'AL_SHAHID_CONTROL'):
        lga.road_quality_idx -= 0.3

    # A2 highway bombs 2043
    if year == 2043 and lga.state in ('Kaduna', 'Niger') and lga.road_quality_idx > 4:
        lga.road_quality_idx -= 1.0

    # --- Market Access ---
    if lga.is_southern:
        lga.market_access_idx += 0.2 * eff
    else:
        lga.market_access_idx += 0.1 * eff

    # Conflict collapses markets
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.market_access_idx = max(1.5, lga.market_access_idx - 0.5)
    elif lga._conflict_status == 'OCCUPIED':
        lga.market_access_idx = max(1.0, lga.market_access_idx - 0.8)

    # Refinery/rail boost
    if lga.refinery_zone:
        lga.market_access_idx += 0.1
    if lga.rail_corridor >= 1:
        lga.market_access_idx += 0.05


def update_education(lga, year):
    """Update education metrics."""
    eff = lga.conflict_effectiveness()

    # Literacy
    if lga.is_southern:
        lga.adult_literacy += 0.7 * eff
        lga.male_literacy += 0.5 * eff
        lga.female_literacy += 0.9 * eff
    else:
        lga.adult_literacy += 0.35 * eff
        lga.male_literacy += 0.3 * eff
        lga.female_literacy += 0.4 * eff

    # Girls' education mandate 2036+
    if year >= 2036 and lga.is_northern:
        lga.female_literacy += 0.3 * eff
        lga.secondary_enrollment += 0.2 * eff
        lga.gender_parity += 0.005 * eff

    # Enrollment
    lga.primary_enrollment += 0.3 * eff
    if lga.is_southern:
        lga.secondary_enrollment += 0.4 * eff
    else:
        lga.secondary_enrollment += 0.2 * eff

    # Out of school children
    if eff > 0.5:
        lga.out_of_school -= 0.3 * eff
    elif lga._conflict_status in ('AL_SHAHID_CONTROL', 'OCCUPIED'):
        lga.out_of_school += 0.5

    # Community college program 2040+
    if year >= 2040:
        lga.secondary_enrollment += 0.15 * eff
        if lga.is_southern and lga.urban_pct > 40:
            lga.tertiary_institution = 1

    # Al-Shahid: education regresses
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.adult_literacy -= 0.2
        lga.female_literacy -= 0.5
        lga.secondary_enrollment -= 0.3
        lga.out_of_school += 1.0
        lga.gender_parity -= 0.01

    # WAFTA cohort returns 2039+
    if year >= 2039 and lga.is_southern and lga.urban_pct > 50:
        lga.adult_literacy += 0.1


def update_economic(lga, year):
    """Update economic indicators: poverty, unemployment, livelihood mix."""
    eff = lga.conflict_effectiveness()

    # --- Poverty reduction ---
    # Accelerating with GDP growth: faster after 2040 when GDP growth accelerates
    gdp_accel = 1.0 + max(0, (year - 2035)) * 0.02  # up to ~1.46 by 2058
    if lga.is_southern and lga.urban_pct > 50:
        lga.poverty_rate -= 1.8 * eff * min(gdp_accel, 1.3)
    elif lga.is_southern:
        lga.poverty_rate -= 0.9 * eff * min(gdp_accel, 1.2)
    elif lga.is_northern and lga.urban_pct > 40:
        lga.poverty_rate -= 0.8 * eff * min(gdp_accel, 1.2)
    elif lga.is_northern:
        lga.poverty_rate -= 0.5 * eff * min(gdp_accel, 1.1)

    # Conflict worsens poverty
    if lga._conflict_status == 'HIGH_CONFLICT':
        lga.poverty_rate += 0.3
    elif lga._conflict_status in ('OCCUPIED', 'AL_SHAHID_CONTROL'):
        lga.poverty_rate += 0.7

    # AI cassava mistake 2039: one-year spike for agricultural LGAs
    if year == 2039 and lga.liv_agriculture > 40:
        lga.poverty_rate += 3.0
    elif year == 2040 and lga.liv_agriculture > 40:
        lga.poverty_rate -= 2.0  # recovery

    # Fuel subsidy phase-out 2044-2046
    if year in (2044, 2045) and lga.is_northern and lga.refinery_zone == 0:
        lga.poverty_rate += 0.5

    # Minimum wage strikes 2045
    if year == 2045:
        lga.unemployment += 2.0
        lga.youth_unemployment += 3.0
    elif year == 2046:
        lga.unemployment -= 1.5
        lga.youth_unemployment -= 2.0

    # Minimum wage raise 2050
    if year == 2050:
        lga.poverty_rate -= 0.5 * eff

    # --- Unemployment ---
    if eff > 0.5:
        lga.unemployment -= 0.3 * eff
        lga.youth_unemployment -= 0.2 * eff

    # Chinese firms leaving 2050-2054
    if 2050 <= year <= 2054 and lga.chinese_economic_presence > 3 and lga.liv_manufacturing > 20:
        lga.unemployment += 0.8
        lga.youth_unemployment += 1.0

    # --- Livelihood mix ---
    # Urbanization shifts from agriculture to services/manufacturing
    if lga.urban_pct > 40:
        shift = 0.3 * eff
        lga.liv_agriculture = max(2, lga.liv_agriculture - shift)
        lga.liv_services += shift * 0.6
        lga.liv_manufacturing += shift * 0.2
        lga.liv_informal -= shift * 0.1

    # Service sector transition post-2045
    if year >= 2045 and lga.urban_pct > 50:
        lga.liv_manufacturing -= 0.15
        lga.liv_services += 0.2
        lga.liv_informal -= 0.05

    # Agricultural mechanization post-2040
    if year >= 2040 and lga.liv_agriculture > 30:
        lga.liv_agriculture -= 0.2 * eff

    # Informal sector decline with BIC
    if lga.bic_effectiveness > 3 and lga.liv_informal > 10:
        lga.liv_informal -= 0.15

    # Refinery effect
    if lga.refinery_zone and year >= 2044:
        lga.liv_extraction += 0.2
        lga.liv_manufacturing += 0.15
        lga.poverty_rate -= 0.3

    lga.normalize_livelihood()


def update_connectivity(lga, year):
    """Update mobile and internet penetration."""
    eff = lga.conflict_effectiveness()

    lga.mobile_penetration += 2.0 * eff
    lga.internet_access += 1.2 * eff

    # 6G towers 2038+
    if year >= 2038 and lga.urban_pct > 40:
        lga.mobile_penetration += 1.5
        lga.internet_access += 1.0

    # Al-Shahid: connectivity stagnates
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.mobile_penetration -= 0.5
        lga.internet_access -= 0.3


def update_bic(lga, year):
    """Update BIC effectiveness."""
    if year < 2032:
        return

    # BIC starts 2032, ramps up
    years_since = year - 2032
    if lga.is_southern and lga.urban_pct > 50:
        lga.bic_effectiveness = min(8.0, years_since * 0.35)
    elif lga.is_southern:
        lga.bic_effectiveness = min(6.0, years_since * 0.2)
    elif lga.is_fct:
        lga.bic_effectiveness = min(7.0, years_since * 0.3)
    elif lga.is_northern and lga._conflict_status == 'STABLE':
        lga.bic_effectiveness = min(4.0, years_since * 0.12)
    elif lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.bic_effectiveness = 0.0
    else:
        lga.bic_effectiveness = min(2.0, years_since * 0.05)

    # Pro-Trust Campaign 2037-2039 boost
    if 2037 <= year <= 2039:
        lga.bic_effectiveness += 0.3

    # Northern police non-cooperation 2046
    if year >= 2046 and lga.is_northern and lga.pct_muslim > 70:
        lga.bic_effectiveness *= 0.8


def update_chinese_presence(lga, year):
    """Update Chinese Economic Presence."""
    if year < 2030:
        # Cobalt prospecting only
        if lga.state == 'Zamfara' and lga.cobalt_extraction_active:
            lga.chinese_economic_presence = 2.0
        return

    if year < 2033:
        if lga.state == 'Zamfara':
            lga.chinese_economic_presence = min(4.0, lga.chinese_economic_presence + 0.5)
        return

    # 2033+ construction workers, WAFTA, joint ventures
    if year <= 2050:
        if lga.is_lagos:
            lga.chinese_economic_presence = min(8.0, lga.chinese_economic_presence + 0.3)
        elif lga.planned_city:
            lga.chinese_economic_presence = min(7.0, lga.chinese_economic_presence + 0.25)
        elif lga.rail_corridor >= 1:
            lga.chinese_economic_presence = min(5.0, lga.chinese_economic_presence + 0.2)
        elif lga.refinery_zone:
            lga.chinese_economic_presence = min(6.0, lga.chinese_economic_presence + 0.2)
        elif lga.is_southern and lga.urban_pct > 50:
            lga.chinese_economic_presence = min(4.0, lga.chinese_economic_presence + 0.15)
        elif lga.cobalt_extraction_active:
            lga.chinese_economic_presence = min(6.0, lga.chinese_economic_presence + 0.2)
    else:
        # 2050+ firms begin leaving
        decline = 0.5 if year < 2054 else 0.8
        lga.chinese_economic_presence = max(1.0, lga.chinese_economic_presence - decline)


def update_cultural(lga, year):
    """Update cultural/linguistic indicators."""
    # English prestige generally rises
    if lga.is_southern:
        lga.english_prestige = min(9.0, lga.english_prestige + 0.05)
    elif lga.is_northern and lga._conflict_status == 'STABLE':
        lga.english_prestige = min(6.0, lga.english_prestige + 0.03)

    # Civilizing Task Force 2051+
    if year >= 2051:
        lga.english_prestige += 0.1

    # Pada areas boost English
    if lga.ethnic[lga.pada_idx] > 2:
        lga.english_prestige = min(9.0, lga.english_prestige + 0.05)

    # Mandarin
    if lga.chinese_economic_presence > 2:
        lga.mandarin_presence = min(7.0, lga.chinese_economic_presence * 0.6)
    if lga.ethnic[lga.naijin_idx] > 1:
        lga.mandarin_presence = max(lga.mandarin_presence, 3.0)
    # Decline with Chinese departure
    if year >= 2050:
        lga.mandarin_presence = max(1.0, lga.mandarin_presence - 0.1)

    # Arabic prestige in north
    if lga.is_northern and lga.pct_muslim > 50:
        if year >= 2040:
            lga.arabic_prestige = min(8.0, lga.arabic_prestige + 0.1)
        else:
            lga.arabic_prestige = min(5.0, lga.arabic_prestige + 0.03)

    # Al-Shahid zones: Arabic very high
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.arabic_prestige = min(9.0, lga.arabic_prestige + 0.3)
        lga.english_prestige = max(2.0, lga.english_prestige - 0.1)


def update_bio_enhancement(lga, year):
    """Update Biological Enhancement Pct based on timeline milestones."""
    if year < 2044:
        return

    # National targets: 10% by 2047, 24% by 2053, 42% by 2057
    # Use higher base progress to hit ~42% national average
    progress = 0.0
    if year <= 2047:
        progress = (year - 2044) / 3.0 * 14.0  # to 14% base
    elif year <= 2053:
        progress = 14.0 + (year - 2047) / 6.0 * 20.0  # to 34%
    elif year <= 2057:
        progress = 34.0 + (year - 2053) / 4.0 * 22.0  # to 56%
    else:
        progress = 56.0

    # Local multiplier (calibrated so national avg ~42% at 2057)
    mult = 1.0
    if lga.is_southern and lga.urban_pct > 60:
        mult = 2.2
    elif lga.is_southern and lga.urban_pct > 30:
        mult = 1.4
    elif lga.is_southern:
        mult = 0.9
    elif lga.is_northern and lga.urban_pct > 50:
        mult = 0.6
    elif lga.is_northern:
        mult = 0.2

    # Wealth and healthcare access factor
    healthcare_factor = lga.access_healthcare / 70.0
    mult *= min(1.5, max(0.3, healthcare_factor))

    # Conflict zones: near zero
    if lga._conflict_status in ('AL_SHAHID_CONTROL', 'OCCUPIED'):
        mult = 0.05

    lga.bio_enhancement_pct = min(85.0, progress * mult)


def update_al_shahid(lga, year):
    """Update al-Shahid influence."""
    if year < 2035:
        return

    if year == 2035:
        # Founded: initial presence in Kano, Kaduna, Bauchi
        if lga.state == 'Kano' and lga.pct_muslim > 70:
            lga.al_shahid_influence = 1.0
        elif lga.state in ('Kaduna', 'Bauchi') and lga.pct_muslim > 60:
            lga.al_shahid_influence = 0.5
        return

    if year <= 2039:
        # Growing slowly
        if lga.state == 'Kano' and lga.pct_muslim > 60:
            lga.al_shahid_influence = min(3.0, lga.al_shahid_influence + 0.4)
        elif lga.is_northern and lga.pct_muslim > 70:
            lga.al_shahid_influence = min(2.0, lga.al_shahid_influence + 0.15)
        # 2036 protests
        if year == 2036 and lga.is_northern and lga.pct_muslim > 60:
            lga.al_shahid_influence += 0.3
        return

    if year == 2040:
        # Kano falls
        if lga.state == 'Kano':
            lga.al_shahid_influence = 4.5
        elif lga.state in ('Jigawa', 'Bauchi', 'Kaduna') and lga.pct_muslim > 60:
            lga.al_shahid_influence = min(3.0, lga.al_shahid_influence + 1.0)
        return

    if 2041 <= year <= 2051:
        if lga.state == 'Kano':
            lga.al_shahid_influence = min(5.0, lga.al_shahid_influence + 0.1)
        elif lga.is_northern and lga.pct_muslim > 70:
            lga.al_shahid_influence = min(3.0, lga.al_shahid_influence + 0.05)
        # 2043 almajiri spread
        if year == 2043 and lga.is_northern:
            lga.al_shahid_influence += 0.2
        # 2046 secret courts
        if year >= 2046 and lga.is_northern and lga.pct_muslim > 80:
            lga.al_shahid_influence += 0.1
        return

    if year >= 2052:
        # Peace talks - decline
        if lga.state == 'Kano':
            lga.al_shahid_influence = max(2.0, lga.al_shahid_influence - 0.4)
        else:
            lga.al_shahid_influence = max(0.0, lga.al_shahid_influence - 0.2)


def update_almajiri(lga, year):
    """Update Almajiri index. Epidemic from 2043."""
    if year < 2043:
        # Slow growth in north
        if lga.is_northern and lga.pct_muslim > 50:
            if lga.almajiri_idx < 3:
                lga.almajiri_idx += 0.05
        return

    # 2043 epidemic
    if lga.is_northern:
        lga.almajiri_idx = min(5, lga.almajiri_idx + 0.2)
    # Spreads to south
    if lga.is_southern and lga.urban_pct > 40:
        lga.almajiri_idx = min(2, lga.almajiri_idx + 0.05)


def update_trad_authority(lga, year):
    """Update traditional authority influence."""
    # Gradual decline with modernization
    if lga.is_southern:
        lga.trad_authority_idx = max(0, lga.trad_authority_idx - 0.02)
    elif lga.is_northern and lga._conflict_status == 'STABLE':
        lga.trad_authority_idx = max(0, lga.trad_authority_idx - 0.01)

    # Sokoto sultanate strengthens during invasion/aftermath
    if lga.state == 'Sokoto' and 2030 <= year <= 2040:
        lga.trad_authority_idx = min(5, lga.trad_authority_idx + 0.05)

    # Al-Shahid areas: traditional authority replaced
    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.trad_authority_idx = max(1, lga.trad_authority_idx - 0.1)


def update_land_formalization(lga, year):
    """Update land formalization percentage."""
    eff = lga.conflict_effectiveness()

    if lga.urban_pct > 50:
        lga.land_formalization += 1.2 * eff
    elif lga.urban_pct > 25:
        lga.land_formalization += 0.5 * eff
    else:
        lga.land_formalization += 0.1 * eff

    if lga._conflict_status == 'AL_SHAHID_CONTROL':
        lga.land_formalization -= 0.3


# ============================================================================
# SECTION 10: SPECIAL EVENT HANDLERS (year-specific)
# ============================================================================

def apply_special_events(lgas, year):
    """Apply year-specific special events that don't fit neatly into the general update functions."""

    if year == 2028:
        # Cobalt discovery in Zamfara
        for lga in lgas:
            if lga.state == 'Zamfara':
                name_l = lga.lga_name.lower()
                if any(s in name_l for s in ['anka', 'gummi', 'bukk', 'maru']):
                    lga.cobalt_extraction_active = 1
                    lga.extraction_intensity = max(lga.extraction_intensity, 3)
            # Oil extraction slight decline
            if lga.oil_producing:
                lga.extraction_intensity = max(0, lga.extraction_intensity - 0.2)

    if year == 2030:
        # ASS invasion damage
        for lga in lgas:
            if lga.state in ASS_INVASION_STATES or lga.state == 'FCT':
                if lga._conflict_status in ('OCCUPIED', 'HIGH_CONFLICT'):
                    lga.road_quality_idx = max(1.0, lga.road_quality_idx - 3.0)
                    lga.market_access_idx = max(1.0, lga.market_access_idx - 3.0)
                    lga.out_of_school += 15.0
                    lga.poverty_rate += 10.0
            # Sokoto: spared but traditional authority rises
            if lga.state == 'Sokoto':
                lga.trad_authority_idx = min(5, lga.trad_authority_idx + 1)

    if year == 2032:
        # Capital moves to Lagos, AZ system
        for lga in lgas:
            if lga.is_lagos:
                lga.road_quality_idx += 1.0
                lga.market_access_idx += 0.5
                lga.access_electricity += 3.0
            # FCT loses capital status
            if lga.is_fct:
                lga.market_access_idx -= 0.3

    if year == 2033:
        # Primary rail corridor construction begins
        for lga in lgas:
            if lga.state in PRIMARY_RAIL_STATES:
                name_l = lga.lga_name.lower()
                # Flag major cities/junction LGAs along the route
                is_on_route = False
                if lga.state == 'Lagos' and lga.urban_pct > 50:
                    is_on_route = True
                elif lga.state == 'Ogun' and any(s in name_l for s in ['abeokuta', 'sagamu', 'ifo']):
                    is_on_route = True
                elif lga.state == 'Oyo' and any(s in name_l for s in ['ibadan', 'oluyole', 'akinyele', 'oyo']):
                    is_on_route = True
                elif lga.state == 'Kwara' and any(s in name_l for s in ['ilorin', 'asa', 'moro']):
                    is_on_route = True
                elif lga.state == 'Niger' and any(s in name_l for s in ['minna', 'chanchaga', 'suleja', 'bida']):
                    is_on_route = True
                elif lga.state == 'FCT':
                    is_on_route = True
                elif lga.state == 'Kaduna' and any(s in name_l for s in ['kaduna', 'chikun', 'igabi', 'zaria']):
                    is_on_route = True
                elif lga.state == 'Kano' and any(s in name_l for s in ['kano', 'nassarawa', 'dala', 'fagge', 'gwale']):
                    is_on_route = True
                elif lga.state == 'Katsina' and any(s in name_l for s in ['katsina', 'funtua']):
                    is_on_route = True

                if is_on_route:
                    lga.rail_corridor = max(lga.rail_corridor, 2)
                elif lga.urban_pct > 40:
                    lga.rail_corridor = max(lga.rail_corridor, 1)

    if year == 2034:
        # Oil peak, BUA pharma, northern rail
        for lga in lgas:
            if lga.oil_producing:
                lga.extraction_intensity = min(5, lga.extraction_intensity + 0.5)

    if year == 2037:
        # AI data center Lagos
        for lga in lgas:
            if lga.is_lagos and lga.urban_pct > 60:
                lga.liv_services += 1.0
                lga.liv_informal -= 0.5
                lga.normalize_livelihood()

    if year == 2039:
        # Eko Atlantic completed
        for lga in lgas:
            if lga.state == 'Lagos':
                name_l = lga.lga_name.lower()
                if any(s in name_l for s in ['eti-osa', 'eti osa', 'lagos island', 'ikoyi']):
                    lga.planned_city = 1
                    lga.housing_affordability = max(2.0, lga.housing_affordability - 2.0)

    if year == 2040:
        # Kano falls - massive population displacement
        for lga in lgas:
            if lga.state == 'Kano':
                lga.bic_effectiveness = 0.0
                lga.market_access_idx = max(1.5, lga.market_access_idx - 3.0)
                lga.out_of_school += 20.0
                lga.poverty_rate += 8.0

    if year == 2043:
        # Almajiri epidemic
        for lga in lgas:
            if lga.is_northern:
                lga.almajiri_idx = min(5, lga.almajiri_idx + 1)
                if lga.almajiri_idx >= 4:
                    lga.out_of_school += 2.0

    if year == 2044:
        # Bonny and Niger-Benue refineries
        for lga in lgas:
            for ref_state, ref_subs, ref_year in REFINERIES:
                if ref_year == 2044 and lga.state == ref_state:
                    name_l = lga.lga_name.lower()
                    if any(s.lower() in name_l for s in ref_subs):
                        lga.refinery_present = 1
                        lga.refinery_zone = 1
                        lga.extraction_intensity = min(5, lga.extraction_intensity + 2)
                        lga.liv_manufacturing += 3
                        lga.liv_extraction += 2
                        lga.normalize_livelihood()

    if year == 2046:
        # Maiduguri refinery (muted by conflict)
        for lga in lgas:
            if lga.state == 'Borno':
                name_l = lga.lga_name.lower()
                if any(s in name_l for s in ['maiduguri', 'jere']):
                    lga.refinery_present = 1
                    lga.refinery_zone = 1
                    lga.extraction_intensity = min(4, lga.extraction_intensity + 1)

    if year == 2048:
        # ASS/EAF immigrants
        for lga in lgas:
            if lga.urban_pct > 30 and lga.urban_pct < 70 and lga.is_southern:
                lga.ethnic[lga.other_idx] += 0.3
                lga.housing_affordability -= 0.2
                lga.normalize_ethnic()

    # Secondary rail expansion (check each year)
    if year in SECONDARY_RAIL:
        branches = SECONDARY_RAIL[year]
        for branch in branches:
            for lga in lgas:
                if lga.state in branch and lga.urban_pct > 20:
                    lga.rail_corridor = max(lga.rail_corridor, 1)

    # Planned cities (check each year for new construction starts)
    for city_name, city_state, city_sub, city_year in PLANNED_CITIES:
        if year == city_year:
            for lga in lgas:
                if lga.state == city_state and city_sub.lower() in lga.lga_name.lower():
                    lga.planned_city = 1


# ============================================================================
# SECTION 11: MAIN SIMULATION LOOP
# ============================================================================

def simulate(lgas, start_year, end_year):
    """Run the year-by-year simulation from start_year to end_year."""
    changelog = []
    changelog.append(f"Simulation: {start_year} -> {end_year}")
    changelog.append(f"LGAs: {len(lgas)}")

    # Snapshot baseline stats
    baseline_pop = sum(l.population for l in lgas)
    baseline_north_poverty = np.mean([l.poverty_rate for l in lgas if l.is_northern])
    baseline_south_poverty = np.mean([l.poverty_rate for l in lgas if l.is_southern])
    baseline_north_literacy = np.mean([l.adult_literacy for l in lgas if l.is_northern])
    baseline_south_literacy = np.mean([l.adult_literacy for l in lgas if l.is_southern])
    changelog.append(f"Baseline population: {baseline_pop:,.0f}")
    changelog.append(f"Baseline North poverty: {baseline_north_poverty:.1f}%")
    changelog.append(f"Baseline South poverty: {baseline_south_poverty:.1f}%")
    changelog.append(f"Baseline North literacy: {baseline_north_literacy:.1f}%")
    changelog.append(f"Baseline South literacy: {baseline_south_literacy:.1f}%")

    for year in range(start_year, end_year + 1):
        print(f"Processing year {year}...")

        # 1. Update conflict status
        for lga in lgas:
            update_conflict_status(lga, year)
            update_federal_control(lga, year)
            update_conflict_history(lga, year)
            lga._development_effectiveness = lga.conflict_effectiveness()

        # 2. Apply special events for this year
        apply_special_events(lgas, year)

        # 3. Update all systems
        for lga in lgas:
            update_fertility(lga, year)
            update_population_growth(lga, year)
            update_age_structure(lga, year)
            update_urbanization(lga, year)
            update_pada_naijin(lga, year)
            update_infrastructure(lga, year)
            update_education(lga, year)
            update_economic(lga, year)
            update_connectivity(lga, year)
            update_bic(lga, year)
            update_chinese_presence(lga, year)
            update_cultural(lga, year)
            update_bio_enhancement(lga, year)
            update_al_shahid(lga, year)
            update_almajiri(lga, year)
            update_trad_authority(lga, year)
            update_land_formalization(lga, year)

        # 4. Migration
        apply_migration(lgas, year)

        # 5. GDP computation
        national_gdp = interpolate_gdp(year)
        for lga in lgas:
            lga.gdp_per_capita = compute_lga_gdp(lga, national_gdp)
            lga.gini_proxy = compute_gini(lga)
            lga.housing_affordability = compute_housing_affordability(lga, year)

        # 6. Clamp all values
        for lga in lgas:
            lga.clamp_percentages()
            lga.normalize_livelihood()

        # Periodic logging
        if year % 5 == 0 or year == end_year:
            total_pop = sum(l.population for l in lgas)
            avg_poverty_n = np.mean([l.poverty_rate for l in lgas if l.is_northern])
            avg_poverty_s = np.mean([l.poverty_rate for l in lgas if l.is_southern])
            avg_lit_n = np.mean([l.adult_literacy for l in lgas if l.is_northern])
            avg_lit_s = np.mean([l.adult_literacy for l in lgas if l.is_southern])
            total_pada = sum(l.population * l.ethnic[l.pada_idx] / 100 for l in lgas)
            total_naijin = sum(l.population * l.ethnic[l.naijin_idx] / 100 for l in lgas)
            national_gdp_now = interpolate_gdp(year)
            print(f"  Year {year}: Pop={total_pop:,.0f}, GDP/cap=${national_gdp_now:,.0f}, "
                  f"Poverty N={avg_poverty_n:.1f}% S={avg_poverty_s:.1f}%, "
                  f"Literacy N={avg_lit_n:.1f}% S={avg_lit_s:.1f}%")
            if year == end_year:
                changelog.append(f"\n=== FINAL STATE {year} ===")
                changelog.append(f"Total population: {total_pop:,.0f}")
                changelog.append(f"National GDP/capita: ${national_gdp_now:,.0f}")
                changelog.append(f"North avg poverty: {avg_poverty_n:.1f}%")
                changelog.append(f"South avg poverty: {avg_poverty_s:.1f}%")
                changelog.append(f"North avg literacy: {avg_lit_n:.1f}%")
                changelog.append(f"South avg literacy: {avg_lit_s:.1f}%")
                changelog.append(f"Total Pada population: {total_pada:,.0f}")
                changelog.append(f"Total Naijin population: {total_naijin:,.0f}")

                # Fertility
                avg_fert_n = np.mean([l.fertility_rate for l in lgas if l.is_northern])
                avg_fert_s = np.mean([l.fertility_rate for l in lgas if l.is_southern])
                changelog.append(f"North avg fertility: {avg_fert_n:.2f}")
                changelog.append(f"South avg fertility: {avg_fert_s:.2f}")

                # Median age
                avg_age_n = np.mean([l.median_age for l in lgas if l.is_northern])
                avg_age_s = np.mean([l.median_age for l in lgas if l.is_southern])
                changelog.append(f"North avg median age: {avg_age_n:.1f}")
                changelog.append(f"South avg median age: {avg_age_s:.1f}")

                # Bio enhancement
                avg_bio = np.mean([l.bio_enhancement_pct for l in lgas])
                changelog.append(f"Avg biological enhancement: {avg_bio:.1f}%")

                # Conflict
                al_shahid_count = sum(1 for l in lgas if l._conflict_status == 'AL_SHAHID_CONTROL')
                stable_count = sum(1 for l in lgas if l._conflict_status == 'STABLE')
                changelog.append(f"LGAs STABLE: {stable_count}, AL_SHAHID_CONTROL: {al_shahid_count}")

    return changelog


# ============================================================================
# SECTION 12: VALIDATION
# ============================================================================

def validate(lgas):
    """Run validation checks on final data. Print results."""
    print("\n=== VALIDATION ===")
    errors = 0
    warnings = 0

    for lga in lgas:
        # 1. Ethnic percentages sum to ~100
        ethnic_total = sum(lga.ethnic)
        if not (95 <= ethnic_total <= 105):
            print(f"  ERROR: {lga.state}/{lga.lga_name} ethnic sum = {ethnic_total:.1f}")
            errors += 1

        # 2. Religious percentages sum to ~100
        relig_total = lga.pct_muslim + lga.pct_christian + lga.pct_traditionalist
        if not (95 <= relig_total <= 105):
            print(f"  ERROR: {lga.state}/{lga.lga_name} religion sum = {relig_total:.1f}")
            errors += 1

        # 3. Livelihood percentages sum to ~100
        liv_total = (lga.liv_agriculture + lga.liv_manufacturing +
                     lga.liv_extraction + lga.liv_services + lga.liv_informal)
        if not (95 <= liv_total <= 105):
            print(f"  ERROR: {lga.state}/{lga.lga_name} livelihood sum = {liv_total:.1f}")
            errors += 1

        # 4. No negative percentages
        for field in ['pct_muslim', 'pct_christian', 'pct_traditionalist',
                       'poverty_rate', 'unemployment', 'youth_unemployment',
                       'adult_literacy', 'access_electricity', 'access_water']:
            if getattr(lga, field) < 0:
                print(f"  ERROR: {lga.state}/{lga.lga_name} {field} = {getattr(lga, field):.1f}")
                errors += 1

        # 5. No percentage > 100
        for field in ['poverty_rate', 'adult_literacy', 'access_electricity',
                       'mobile_penetration', 'internet_access']:
            if getattr(lga, field) > 100:
                print(f"  WARN: {lga.state}/{lga.lga_name} {field} = {getattr(lga, field):.1f} (capped)")
                setattr(lga, field, 100.0)
                warnings += 1

        # 6. Population positive
        if lga.population <= 0:
            print(f"  ERROR: {lga.state}/{lga.lga_name} population = {lga.population}")
            errors += 1

        # 8. Age consistency
        if lga.median_age > 30 and lga.pct_under_30 > 60:
            warnings += 1
        if lga.median_age < 22 and lga.pct_under_30 < 50:
            warnings += 1

        # 9. Fertility range
        if not (1.0 <= lga.fertility_rate <= 5.0):
            print(f"  WARN: {lga.state}/{lga.lga_name} fertility = {lga.fertility_rate:.2f}")
            warnings += 1

    # 7. Total national population
    total_pop = sum(l.population for l in lgas)
    if not (250_000_000 <= total_pop <= 350_000_000):
        print(f"  WARN: National population = {total_pop:,.0f} (outside 250M-350M range)")
        warnings += 1
    else:
        print(f"  OK: National population = {total_pop:,.0f}")

    print(f"\nValidation complete: {errors} errors, {warnings} warnings")
    return errors, warnings


# ============================================================================
# SECTION 13: OUTPUT COLUMN DEFINITIONS
# ============================================================================

# Define the output column structure
# Each entry: (group_header, column_name, data_type, description, getter_func)

def build_output_columns():
    """Return list of output column definitions."""
    cols = []

    def add(group, name, dtype, desc, getter):
        cols.append((group, name, dtype, desc, getter))

    # IDENTIFICATION
    add('IDENTIFICATION', 'State', 'text', 'Nigerian state name', lambda l: l.state)
    add('IDENTIFICATION', 'LGA Name', 'text', 'Local Government Area name', lambda l: l.lga_name)
    add('IDENTIFICATION', 'Colonial Era Region', 'text', 'Historical colonial region', lambda l: l.colonial_region)
    add('IDENTIFICATION', 'Terrain Type', 'text', 'Geographic terrain classification', lambda l: l.terrain)

    # ADMINISTRATIVE (NEW)
    add('ADMINISTRATIVE', 'Administrative Zone', 'integer', 'Administrative Zone number (1-8), established 2032', lambda l: l.admin_zone)
    add('ADMINISTRATIVE', 'AZ Name', 'text', 'Administrative Zone name', lambda l: l.az_name)

    # DEMOGRAPHIC
    add('DEMOGRAPHIC', 'Estimated Population', 'integer', 'Estimated total population 2058', lambda l: round(l.population))
    add('DEMOGRAPHIC', 'Population Density per km2', 'float', 'Population per square kilometer', lambda l: round(l.pop_density, 1))
    add('DEMOGRAPHIC', 'Median Age Estimate', 'float', 'Estimated median age', lambda l: round(l.median_age, 1))
    add('DEMOGRAPHIC', '% Population Under 30', 'float', 'Percentage of population under 30', lambda l: round(l.pct_under_30, 1))
    add('DEMOGRAPHIC', 'Fertility Rate Est', 'float', 'Estimated total fertility rate', lambda l: round(l.fertility_rate, 2))
    add('DEMOGRAPHIC', 'Biological Enhancement Pct', 'float', 'Percentage of population biologically enhanced', lambda l: round(l.bio_enhancement_pct, 1))

    # ETHNOLINGUISTIC (86 ethnic columns)
    for i, name in enumerate(ETHNIC_COLS_NAMES):
        idx = i  # index into lga.ethnic array
        add('ETHNOLINGUISTIC COMPOSITION', f'% {name}', 'float',
            f'Percentage of population identifying as {name}',
            lambda l, idx=idx: round(l.ethnic[idx], 2))

    # RELIGIOUS COMPOSITION
    add('RELIGIOUS COMPOSITION', '% Muslim', 'float', 'Percentage Muslim', lambda l: round(l.pct_muslim, 1))
    add('RELIGIOUS COMPOSITION', '% Christian', 'float', 'Percentage Christian', lambda l: round(l.pct_christian, 1))
    add('RELIGIOUS COMPOSITION', '% Traditionalist', 'float', 'Percentage Traditionalist', lambda l: round(l.pct_traditionalist, 1))
    add('RELIGIOUS COMPOSITION', 'Tijaniyya Presence', 'integer', 'Tijaniyya order presence (0-3)', lambda l: l.tijaniyya)
    add('RELIGIOUS COMPOSITION', 'Qadiriyya Presence', 'integer', 'Qadiriyya order presence (0-3)', lambda l: l.qadiriyya)
    add('RELIGIOUS COMPOSITION', 'Pentecostal Growth', 'integer', 'Pentecostal Christianity growth (0-3)', lambda l: l.pentecostal)
    add('RELIGIOUS COMPOSITION', 'Traditionalist Practice', 'integer', 'Traditional religious practice persistence (0-3)', lambda l: l.trad_practice)
    add('RELIGIOUS COMPOSITION', 'Al-Shahid Influence', 'float', 'Al-Shahid movement influence (0-5)', lambda l: round(l.al_shahid_influence, 1))
    add('RELIGIOUS COMPOSITION', 'Almajiri Index', 'integer', 'Almajiri prevalence index (0-5)', lambda l: min(5, max(0, round(l.almajiri_idx))))
    add('RELIGIOUS COMPOSITION', 'Religious Notes', 'text', 'Qualitative religious composition notes', lambda l: l.religious_notes)

    # URBANIZATION
    add('URBANIZATION', 'Major Urban Center', 'text', 'Primary urban center in LGA', lambda l: l.major_urban_center)
    add('URBANIZATION', 'Urban Pct', 'float', 'Percentage urban population', lambda l: round(l.urban_pct, 1))

    # ECONOMIC
    add('ECONOMIC', 'Oil Producing', 'integer', 'Oil producing LGA (0/1)', lambda l: l.oil_producing)
    add('ECONOMIC', 'Pct Livelihood Agriculture', 'float', 'Agriculture share of livelihoods', lambda l: round(l.liv_agriculture, 1))
    add('ECONOMIC', 'Pct Livelihood Manufacturing', 'float', 'Manufacturing share of livelihoods', lambda l: round(l.liv_manufacturing, 1))
    add('ECONOMIC', 'Pct Livelihood Extraction', 'float', 'Resource extraction share', lambda l: round(l.liv_extraction, 1))
    add('ECONOMIC', 'Pct Livelihood Services', 'float', 'Services sector share', lambda l: round(l.liv_services, 1))
    add('ECONOMIC', 'Pct Livelihood Informal', 'float', 'Informal sector share', lambda l: round(l.liv_informal, 1))
    add('ECONOMIC', 'Livelihood Notes', 'text', 'Original livelihood description (2026)', lambda l: l.livelihood_notes)
    add('ECONOMIC', 'Oil Extraction Active', 'integer', 'Active oil/gas extraction (0/1)', lambda l: l.oil_extraction_active)
    add('ECONOMIC', 'Cobalt Extraction Active', 'integer', 'Active cobalt mining (0/1)', lambda l: l.cobalt_extraction_active)
    add('ECONOMIC', 'Other Mining Active', 'integer', 'Other mining activity (0/1)', lambda l: l.other_mining_active)
    add('ECONOMIC', 'Refinery Present', 'integer', 'Oil refinery in LGA (0/1)', lambda l: l.refinery_present)
    add('ECONOMIC', 'Extraction Intensity', 'integer', 'Resource extraction intensity (0-5)', lambda l: min(5, max(0, round(l.extraction_intensity))))
    add('ECONOMIC', 'Poverty Rate Pct', 'float', 'Poverty rate percentage', lambda l: round(l.poverty_rate, 1))
    add('ECONOMIC', 'Unemployment Rate Pct', 'float', 'General unemployment rate', lambda l: round(l.unemployment, 1))
    add('ECONOMIC', 'Youth Unemployment Rate Pct', 'float', 'Youth unemployment rate', lambda l: round(l.youth_unemployment, 1))
    add('ECONOMIC', 'GDP Per Capita Est', 'float', 'Estimated GDP per capita (USD)', lambda l: round(l.gdp_per_capita, 0))
    add('ECONOMIC', 'Gini Proxy', 'float', 'Local inequality estimate (0-1)', lambda l: round(l.gini_proxy, 3))
    add('ECONOMIC', 'Chinese Economic Presence', 'float', 'Chinese economic activity index (0-10)', lambda l: round(l.chinese_economic_presence, 1))
    add('ECONOMIC', 'Rail Corridor', 'integer', 'Rail access: 0=none, 1=secondary, 2=primary HSR', lambda l: l.rail_corridor)
    add('ECONOMIC', 'Planned City', 'integer', 'Contains or adjacent to planned city (0/1)', lambda l: l.planned_city)
    add('ECONOMIC', 'Refinery Zone', 'integer', 'Near refinery complex (0/1)', lambda l: l.refinery_zone)
    add('ECONOMIC', 'Housing Affordability', 'float', 'Housing affordability index (0-10, 10=affordable)', lambda l: round(l.housing_affordability, 1))

    # INFRASTRUCTURE
    add('INFRASTRUCTURE', 'Access Electricity Pct', 'float', 'Electricity access percentage', lambda l: round(l.access_electricity, 1))
    add('INFRASTRUCTURE', 'Access Water Pct', 'float', 'Clean water access percentage', lambda l: round(l.access_water, 1))
    add('INFRASTRUCTURE', 'Access Healthcare Pct', 'float', 'Healthcare access percentage', lambda l: round(l.access_healthcare, 1))
    add('INFRASTRUCTURE', 'Road Quality Index', 'float', 'Road infrastructure quality (1-10)', lambda l: round(l.road_quality_idx, 1))
    add('INFRASTRUCTURE', 'Road Quality Notes', 'text', 'Original road quality description (2026)', lambda l: str(l.road_quality_raw) if l.road_quality_raw else '')
    add('INFRASTRUCTURE', 'Market Access Index', 'float', 'Market access quality (1-10)', lambda l: round(l.market_access_idx, 1))
    add('INFRASTRUCTURE', 'Market Access Notes', 'text', 'Original market access description (2026)', lambda l: str(l.market_access_raw) if l.market_access_raw else '')

    # EDUCATION
    add('EDUCATION', 'Adult Literacy Rate Pct', 'float', 'Adult literacy rate', lambda l: round(l.adult_literacy, 1))
    add('EDUCATION', 'Male Literacy Rate Pct', 'float', 'Male literacy rate', lambda l: round(l.male_literacy, 1))
    add('EDUCATION', 'Female Literacy Rate Pct', 'float', 'Female literacy rate', lambda l: round(l.female_literacy, 1))
    add('EDUCATION', 'Primary Enrollment Pct', 'float', 'Primary school enrollment rate', lambda l: round(l.primary_enrollment, 1))
    add('EDUCATION', 'Secondary Enrollment Pct', 'float', 'Secondary school enrollment rate', lambda l: round(l.secondary_enrollment, 1))
    add('EDUCATION', 'Gender Parity Index', 'float', 'Education gender parity index', lambda l: round(l.gender_parity, 3))
    add('EDUCATION', 'Out of School Children Pct', 'float', 'Out of school children percentage', lambda l: round(l.out_of_school, 1))
    add('EDUCATION', 'Num Secondary Schools', 'integer', 'Number of secondary schools', lambda l: max(0, round(l.num_secondary_schools)))
    add('EDUCATION', 'Tertiary Institution', 'integer', 'Tertiary institution present (0/1)', lambda l: l.tertiary_institution)

    # POLITICAL STRUCTURE
    add('POLITICAL STRUCTURE', 'Traditional Authority', 'text', 'Traditional ruler/authority name', lambda l: l.trad_authority)
    add('POLITICAL STRUCTURE', 'Trad Authority Index', 'integer', 'Traditional authority influence (0-5)', lambda l: min(5, max(0, round(l.trad_authority_idx))))
    add('POLITICAL STRUCTURE', 'Trad Authority Notes', 'text', 'Traditional authority qualitative notes', lambda l: l.trad_authority_notes)
    add('POLITICAL STRUCTURE', 'Land Formalization Pct', 'float', 'Land tenure formalization percentage', lambda l: round(l.land_formalization, 1))
    add('POLITICAL STRUCTURE', 'Land Tenure Notes', 'text', 'Original land tenure description (2026)', lambda l: l.land_tenure_notes)
    add('POLITICAL STRUCTURE', 'Conflict History', 'text', 'Worst conflict status experienced', lambda l: l.conflict_history)
    add('POLITICAL STRUCTURE', 'Federal Control 2058', 'text', 'Federal control status at 2058', lambda l: l.federal_control)
    add('POLITICAL STRUCTURE', 'BIC Effectiveness', 'float', 'BIC anti-corruption effectiveness (0-10)', lambda l: round(l.bic_effectiveness, 1))

    # CONNECTIVITY
    add('CONNECTIVITY', 'Mobile Phone Penetration Pct', 'float', 'Mobile phone penetration', lambda l: round(l.mobile_penetration, 1))
    add('CONNECTIVITY', 'Internet Access Pct', 'float', 'Internet access percentage', lambda l: round(l.internet_access, 1))

    # CULTURAL (NEW)
    add('CULTURAL', 'English Prestige', 'float', 'English language prestige (0-10)', lambda l: round(l.english_prestige, 1))
    add('CULTURAL', 'Mandarin Presence', 'float', 'Mandarin language presence (0-10)', lambda l: round(l.mandarin_presence, 1))
    add('CULTURAL', 'Arabic Prestige', 'float', 'Arabic language prestige (0-10)', lambda l: round(l.arabic_prestige, 1))

    # DATA QUALITY
    add('DATA QUALITY', 'Data Notes', 'text', 'Data quality notes', lambda l: l.data_notes)

    return cols


# ============================================================================
# SECTION 14: WRITE OUTPUT
# ============================================================================

def write_output(lgas, changelog, output_file):
    """Write the output Excel file."""
    print(f"\nWriting output to {output_file}...")

    wb = Workbook()

    # --- Sheet 1: LGA_DATA ---
    ws = wb.active
    ws.title = 'LGA_DATA'

    output_cols = build_output_columns()

    # Row 1: Group headers
    current_group = None
    for col_idx, (group, name, dtype, desc, getter) in enumerate(output_cols):
        if group != current_group:
            ws.cell(row=1, column=col_idx + 1, value=group)
            current_group = group

    # Row 2: Column names
    for col_idx, (group, name, dtype, desc, getter) in enumerate(output_cols):
        ws.cell(row=2, column=col_idx + 1, value=name)

    # Data rows
    for row_idx, lga in enumerate(lgas):
        for col_idx, (group, name, dtype, desc, getter) in enumerate(output_cols):
            try:
                val = getter(lga)
            except Exception:
                val = None
            ws.cell(row=row_idx + 3, column=col_idx + 1, value=val)

    print(f"  Wrote {len(lgas)} LGA rows with {len(output_cols)} columns")

    # --- Sheet 2: METADATA ---
    ws_meta = wb.create_sheet('METADATA')
    ws_meta.cell(row=1, column=1, value='Column Letter')
    ws_meta.cell(row=1, column=2, value='Column Name')
    ws_meta.cell(row=1, column=3, value='Data Type')
    ws_meta.cell(row=1, column=4, value='Group')
    ws_meta.cell(row=1, column=5, value='Description')

    for i, (group, name, dtype, desc, getter) in enumerate(output_cols):
        col_letter = get_column_letter(i + 1)
        ws_meta.cell(row=i + 2, column=1, value=col_letter)
        ws_meta.cell(row=i + 2, column=2, value=name)
        ws_meta.cell(row=i + 2, column=3, value=dtype)
        ws_meta.cell(row=i + 2, column=4, value=group)
        ws_meta.cell(row=i + 2, column=5, value=desc)

    print(f"  Wrote METADATA with {len(output_cols)} column entries")

    # --- Sheet 3: CHANGELOG ---
    ws_log = wb.create_sheet('CHANGELOG')
    ws_log.cell(row=1, column=1, value='Simulation Changelog')
    for i, line in enumerate(changelog):
        ws_log.cell(row=i + 2, column=1, value=line)

    # Add event summary
    events = [
        "2028: Cobalt discovery in Zamfara; ISWAP escalation; banditry rise",
        "2029: Diaspora return begins; tiny pada trickle to Lagos",
        "2030: ASS invasion from Sahel; Zamfara occupied; corridor to Abuja",
        "2031: Liberation; ASS expelled; post-war displacement",
        "2032: Fifth Republic; capital to Lagos; AZ system; BIC established; pada arrive",
        "2033: WAFTA education pipeline; primary rail construction begins",
        "2034: Oil peak; BUA pharma; northern rail expansion under Danjuma",
        "2035: al-Shahid founded; New Anka planned city; IDA Corp",
        "2036: Girls education mandate; al-Shahid protests",
        "2037: Road improvement program; AI data center; Pro-Trust Campaign; New Makoko",
        "2038: Southern dominance crystallizes; Naijin emerge; 6G towers",
        "2039: WAFTA cohort returns; AI cassava error; Eko Atlantic complete",
        "2040: Lagos bombing; anti-Muslim pogroms; KANO FALLS TO AL-SHAHID",
        "2041-42: Al-Shahid war; GDP growth continues in south",
        "2043: Almajiri epidemic; A2 highway bombs",
        "2044: Bonny Island + Niger-Benue refineries operational",
        "2045: Minimum wage strikes; automation begins; Shiroro dam attack",
        "2046: Sharia crisis; Maiduguri refinery; northern police non-cooperation",
        "2047: GDP $7k; 10% biologically enhanced",
        "2048: ASS/EAF immigrants arrive; Lagos housing crisis",
        "2050: Chinese firms begin leaving; minimum wage raise",
        "2051: Civilizing Task Force; English mandate",
        "2052: GDP $13k; al-Shahid peace talks begin",
        "2053: 24% biologically enhanced; women's rights movement",
        "2054: Chinese departure accelerates",
        "2057: GDP $26k; 42% biologically enhanced",
        "2058: Khalil steps down; game start snapshot",
    ]
    start_row = len(changelog) + 4
    ws_log.cell(row=start_row, column=1, value="=== EVENT TIMELINE APPLIED ===")
    for i, evt in enumerate(events):
        ws_log.cell(row=start_row + 1 + i, column=1, value=evt)

    wb.save(output_file)
    print(f"  Saved {output_file}")


# ============================================================================
# SECTION 15: MAIN
# ============================================================================

def main():
    print("=" * 60)
    print("Nigeria PolSim: 2026 -> 2058 LGA Projection")
    print("=" * 60)

    # Load data
    raw_rows, col_names, group_headers, metadata = load_data(INPUT_FILE)

    # Create LGA objects
    print("Initializing LGA objects...")
    lgas = []
    for row in raw_rows:
        lga = LGA(row, col_names)
        lgas.append(lga)
    print(f"  Initialized {len(lgas)} LGAs")

    # Verify AZ assignment
    az_counts = {}
    for lga in lgas:
        az_counts[lga.admin_zone] = az_counts.get(lga.admin_zone, 0) + 1
    print("  AZ distribution:", dict(sorted(az_counts.items())))

    # Initial normalization
    for lga in lgas:
        lga.normalize_ethnic()
        lga.normalize_religion()
        lga.normalize_livelihood()
        lga.clamp_percentages()

    # Run simulation
    changelog = simulate(lgas, START_YEAR, END_YEAR)

    # Validate
    errors, warnings = validate(lgas)

    # Write output
    write_output(lgas, changelog, OUTPUT_FILE)

    print("\n" + "=" * 60)
    print("DONE.")
    print("=" * 60)


if __name__ == '__main__':
    main()
