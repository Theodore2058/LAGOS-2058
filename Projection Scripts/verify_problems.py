#!/usr/bin/env python3
"""Verify all reported problems in the spreadsheet."""
import numpy as np
from openpyxl import load_workbook
from collections import defaultdict

FILE = "nigeria_lga_polsim_2058.xlsx"

NORTHERN_STATES = {
    'Adamawa', 'Bauchi', 'Benue', 'Borno', 'FCT', 'Gombe', 'Jigawa',
    'Kaduna', 'Kano', 'Katsina', 'Kebbi', 'Kogi', 'Kwara', 'Nasarawa',
    'Niger', 'Plateau', 'Sokoto', 'Taraba', 'Yobe', 'Zamfara'
}
SOUTHERN_STATES = {
    'Abia', 'Akwa Ibom', 'Anambra', 'Bayelsa', 'Cross River', 'Delta',
    'Ebonyi', 'Edo', 'Ekiti', 'Enugu', 'Imo', 'Lagos', 'Ogun', 'Ondo',
    'Osun', 'Oyo', 'Rivers'
}

def sf(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except: return default

def main():
    wb = load_workbook(FILE, data_only=True)
    ws = wb['LGA_DATA']

    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

    def ci(name):
        try: return headers.index(name)
        except ValueError: return None

    # Print all headers for reference
    print("=== COLUMN HEADERS ===")
    for i, h in enumerate(headers):
        if h:
            print(f"  {i}: {h}")

    data = []
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if row[0] is not None:
            data.append(row)

    print(f"\nTotal LGAs: {len(data)}")

    # Identify key columns
    state_col = ci('State')
    gdp_col = ci('GDP Per Capita Est')
    poverty_col = ci('Poverty Rate Pct')
    pop_col = ci('Estimated Population')
    literacy_col = ci('Adult Literacy Rate Pct')
    fertility_col = ci('Fertility Rate Est')
    shahid_col = ci('Al-Shahid Influence')
    urban_col = ci('Urban Pct')
    road_col = ci('Road Quality Index')
    cobalt_col = ci('Cobalt Extraction Active')
    refinery_col = ci('Refinery Present')
    conflict_col = ci('Conflict History')
    almajiri_col = ci('Almajiri Index')
    bio_col = ci('Biological Enhancement Pct')
    pada_col = ci('% Pada')
    naijin_col = ci('% Naijin')
    lga_col = ci('LGA Name')

    print(f"\n=== KEY COLUMN INDICES ===")
    for name, idx in [('State', state_col), ('GDP', gdp_col), ('Poverty', poverty_col),
                       ('Population', pop_col), ('Literacy', literacy_col), ('Fertility', fertility_col),
                       ('Al-Shahid', shahid_col), ('Urban', urban_col), ('Road', road_col),
                       ('Cobalt', cobalt_col), ('Refinery', refinery_col), ('Conflict', conflict_col),
                       ('Almajiri', almajiri_col), ('Bio Enhancement', bio_col),
                       ('Pada', pada_col), ('Naijin', naijin_col), ('LGA', lga_col)]:
        print(f"  {name}: {idx}")

    # ==========================================
    # 1. GDP Per Capita
    # ==========================================
    print("\n" + "="*60)
    print("1. GDP PER CAPITA CHECK")
    print("="*60)
    gdps = [(sf(r[gdp_col]), sf(r[pop_col])) for r in data if gdp_col and pop_col]
    total_gdp_weighted = sum(g*p for g,p in gdps)
    total_pop = sum(p for g,p in gdps)
    weighted_avg = total_gdp_weighted / total_pop if total_pop > 0 else 0
    print(f"  Weighted national avg GDP/capita: ${weighted_avg:,.0f}")
    print(f"  Target: ~$26,000-$28,000")

    # Lagos avg
    lagos_gdps = [(sf(r[gdp_col]), sf(r[pop_col])) for r in data if r[state_col] == 'Lagos']
    if lagos_gdps:
        lagos_avg = sum(g*p for g,p in lagos_gdps) / sum(p for g,p in lagos_gdps)
        print(f"  Lagos weighted avg: ${lagos_avg:,.0f}")

    # North vs South
    north_gdps = [(sf(r[gdp_col]), sf(r[pop_col])) for r in data if r[state_col] in NORTHERN_STATES]
    south_gdps = [(sf(r[gdp_col]), sf(r[pop_col])) for r in data if r[state_col] in SOUTHERN_STATES]
    if north_gdps:
        n_avg = sum(g*p for g,p in north_gdps) / sum(p for g,p in north_gdps)
        print(f"  North weighted avg: ${n_avg:,.0f}")
    if south_gdps:
        s_avg = sum(g*p for g,p in south_gdps) / sum(p for g,p in south_gdps)
        print(f"  South weighted avg: ${s_avg:,.0f}")

    # ==========================================
    # 2. Northern Poverty
    # ==========================================
    print("\n" + "="*60)
    print("2. NORTHERN POVERTY CHECK")
    print("="*60)
    north_pov = [sf(r[poverty_col]) for r in data if r[state_col] in NORTHERN_STATES and poverty_col]
    south_pov = [sf(r[poverty_col]) for r in data if r[state_col] in SOUTHERN_STATES and poverty_col]
    print(f"  North mean poverty: {np.mean(north_pov):.1f}%  (target: 30-45%)")
    print(f"  South mean poverty: {np.mean(south_pov):.1f}%")

    # ==========================================
    # 3. Northern Literacy
    # ==========================================
    print("\n" + "="*60)
    print("3. NORTHERN LITERACY CHECK")
    print("="*60)
    north_lit = [sf(r[literacy_col]) for r in data if r[state_col] in NORTHERN_STATES and literacy_col]
    south_lit = [sf(r[literacy_col]) for r in data if r[state_col] in SOUTHERN_STATES and literacy_col]
    print(f"  North mean literacy: {np.mean(north_lit):.1f}%  (target: 65-75%)")
    print(f"  South mean literacy: {np.mean(south_lit):.1f}%")

    # ==========================================
    # 4. Total Population
    # ==========================================
    print("\n" + "="*60)
    print("4. TOTAL POPULATION CHECK")
    print("="*60)
    total = sum(sf(r[pop_col]) for r in data if pop_col) / 1e6
    print(f"  Total population: {total:.1f}M  (target: 280-320M)")

    north_pop = sum(sf(r[pop_col]) for r in data if r[state_col] in NORTHERN_STATES) / 1e6
    south_pop = sum(sf(r[pop_col]) for r in data if r[state_col] in SOUTHERN_STATES) / 1e6
    print(f"  North: {north_pop:.1f}M, South: {south_pop:.1f}M")

    # ==========================================
    # 5. Zero Poverty LGAs
    # ==========================================
    print("\n" + "="*60)
    print("5. ZERO POVERTY LGAs CHECK")
    print("="*60)
    zero_pov = [r for r in data if sf(r[poverty_col]) == 0]
    near_zero = [r for r in data if 0 <= sf(r[poverty_col]) <= 1]
    print(f"  LGAs at exactly 0% poverty: {len(zero_pov)}")
    print(f"  LGAs at 0-1% poverty: {len(near_zero)}")
    if zero_pov:
        states_with_zero = defaultdict(int)
        for r in zero_pov:
            states_with_zero[r[state_col]] += 1
        for s, c in sorted(states_with_zero.items(), key=lambda x: -x[1])[:10]:
            print(f"    {s}: {c}")

    # ==========================================
    # 6. Fertility Floor
    # ==========================================
    print("\n" + "="*60)
    print("6. FERTILITY FLOOR CHECK")
    print("="*60)
    below_13 = [(r[state_col], r[lga_col], sf(r[fertility_col])) for r in data if sf(r[fertility_col]) < 1.3]
    print(f"  LGAs below 1.3 fertility: {len(below_13)}")
    if below_13:
        below_13.sort(key=lambda x: x[2])
        for s, l, f in below_13[:10]:
            print(f"    {s}-{l}: {f:.2f}")

    # ==========================================
    # 7. Al-Shahid in Kano
    # ==========================================
    print("\n" + "="*60)
    print("7. AL-SHAHID INFLUENCE IN KANO CHECK")
    print("="*60)
    if shahid_col:
        kano_shahid = [sf(r[shahid_col]) for r in data if r[state_col] == 'Kano']
        print(f"  Kano mean: {np.mean(kano_shahid):.2f}  (target: 3.5-4.5)")
        print(f"  Kano max: {max(kano_shahid):.2f}")
        print(f"  Kano min: {min(kano_shahid):.2f}")

    # ==========================================
    # 8. Lagos Urbanization
    # ==========================================
    print("\n" + "="*60)
    print("8. LAGOS URBANIZATION CHECK")
    print("="*60)
    if urban_col:
        lagos_urban = [(r[lga_col], sf(r[urban_col])) for r in data if r[state_col] == 'Lagos']
        lagos_urban.sort(key=lambda x: x[1])
        print(f"  Lagos LGAs sorted by urban %:")
        for l, u in lagos_urban[:10]:
            print(f"    {l}: {u:.1f}%")
        print(f"  Lagos min urban: {min(u for l,u in lagos_urban):.1f}%")
        print(f"  Lagos mean urban: {np.mean([u for l,u in lagos_urban]):.1f}%")

    # ==========================================
    # 9. Southern Road Quality
    # ==========================================
    print("\n" + "="*60)
    print("9. SOUTHERN ROAD QUALITY CHECK")
    print("="*60)
    if road_col:
        south_roads = [sf(r[road_col]) for r in data if r[state_col] in SOUTHERN_STATES]
        print(f"  South road quality: min={min(south_roads):.1f}, max={max(south_roads):.1f}, mean={np.mean(south_roads):.1f}, std={np.std(south_roads):.2f}")
        print(f"  Target range: 5-6 (remote) to 10 (highways)")

    # ==========================================
    # 10. Cobalt in Katsina
    # ==========================================
    print("\n" + "="*60)
    print("10. COBALT IN KATSINA CHECK")
    print("="*60)
    if cobalt_col:
        cobalt_lgas = [(r[state_col], r[lga_col], sf(r[cobalt_col])) for r in data if sf(r[cobalt_col]) > 0]
        print(f"  LGAs with cobalt extraction active:")
        for s, l, c in cobalt_lgas:
            print(f"    {s}-{l}: {c}")
        katsina_cobalt = [r for r in data if r[state_col] == 'Katsina' and sf(r[cobalt_col]) > 0]
        print(f"  Katsina LGAs with cobalt: {len(katsina_cobalt)}")

    # ==========================================
    # 11. Unauthorized Refineries
    # ==========================================
    print("\n" + "="*60)
    print("11. UNAUTHORIZED REFINERIES CHECK")
    print("="*60)
    if refinery_col:
        refinery_lgas = [(r[state_col], r[lga_col], sf(r[refinery_col])) for r in data if sf(r[refinery_col]) > 0]
        print(f"  LGAs with refinery present:")
        for s, l, c in refinery_lgas:
            print(f"    {s}-{l}: {c}")

    # ==========================================
    # 12. MODERATE Conflict Category
    # ==========================================
    print("\n" + "="*60)
    print("12. CONFLICT CATEGORIES CHECK")
    print("="*60)
    if conflict_col:
        categories = defaultdict(int)
        for r in data:
            cat = r[conflict_col]
            categories[cat] += 1
        print(f"  Conflict categories found:")
        for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
            print(f"    {cat}: {count}")

    # ==========================================
    # 13. Almajiri in South
    # ==========================================
    print("\n" + "="*60)
    print("13. ALMAJIRI IN SOUTH CHECK")
    print("="*60)
    if almajiri_col:
        south_almajiri_nonzero = [(r[state_col], r[lga_col], sf(r[almajiri_col]))
                                   for r in data if r[state_col] in SOUTHERN_STATES and sf(r[almajiri_col]) > 0]
        print(f"  Southern LGAs with Almajiri > 0: {len(south_almajiri_nonzero)}")

        almajiri_5 = [r for r in data if sf(r[almajiri_col]) == 5]
        print(f"  Total LGAs at Almajiri level 5: {len(almajiri_5)}")

        # Distribution
        for level in range(6):
            count = sum(1 for r in data if int(sf(r[almajiri_col])) == level)
            print(f"    Level {level}: {count}")

    # ==========================================
    # 14. Biological Enhancement
    # ==========================================
    print("\n" + "="*60)
    print("14. BIOLOGICAL ENHANCEMENT CHECK")
    print("="*60)
    if bio_col and pop_col:
        bio_vals = [(sf(r[bio_col]), sf(r[pop_col])) for r in data]
        weighted_bio = sum(b*p for b,p in bio_vals) / sum(p for b,p in bio_vals)
        print(f"  Weighted national avg: {weighted_bio:.1f}%  (target: 42%)")

    # ==========================================
    # 15. Pada Distribution in Lagos
    # ==========================================
    print("\n" + "="*60)
    print("15. PADA DISTRIBUTION IN LAGOS CHECK")
    print("="*60)
    if pada_col:
        lagos_pada = [(r[lga_col], sf(r[pada_col])) for r in data if r[state_col] == 'Lagos']
        lagos_pada.sort(key=lambda x: -x[1])
        print(f"  Lagos Pada distribution:")
        for l, p in lagos_pada:
            print(f"    {l}: {p:.2f}%")

    wb.close()

if __name__ == '__main__':
    main()
