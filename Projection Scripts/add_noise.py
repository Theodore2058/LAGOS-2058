#!/usr/bin/env python3
"""
Post-processing script: adds Gaussian noise to nigeria_lga_polsim_2058.xlsx
to diversify within-state values that are too uniform.

Noise is:
  - Proportional to the column's plausible range (not raw value)
  - Modulated by LGA-level factors (urban/rural, population, terrain)
  - Bounded to respect column min/max
  - Deterministic (fixed seed)
  - Preserves normalization constraints (ethnic/religious/livelihood sums)
"""

import math
import numpy as np
from openpyxl import load_workbook
from copy import deepcopy

INPUT_FILE = "nigeria_lga_polsim_2058.xlsx"
OUTPUT_FILE = "nigeria_lga_polsim_2058.xlsx"  # overwrite in place
SEED = 7777
rng = np.random.default_rng(SEED)

# ============================================================================
# COLUMN NOISE PROFILES
# ============================================================================
# Each entry: column_name -> (noise_std_fraction, min_val, max_val, noise_type)
#   noise_std_fraction: std of noise as fraction of (max_val - min_val)
#   noise_type:
#     'proportional' = noise scaled by current value (louder where value is higher)
#     'absolute'     = fixed std regardless of value
#     'urban_mod'    = noise modulated by urban/rural split (more variation in diverse states)

# --- HIGH priority: state-level paste columns ---
HIGH_NOISE = {
    'Chinese Economic Presence': (0.12, 0.0, 10.0, 'urban_mod'),
    'Mandarin Presence':         (0.12, 0.0, 10.0, 'urban_mod'),
    'Arabic Prestige':           (0.10, 0.0, 10.0, 'urban_mod'),
    'English Prestige':          (0.08, 0.0, 10.0, 'urban_mod'),
    'Al-Shahid Influence':       (0.10, 0.0, 5.0,  'urban_mod'),
    'Almajiri Index':            (0.12, 0,   5,     'urban_mod'),
    'BIC Effectiveness':         (0.10, 0.0, 10.0,  'urban_mod'),
}

# --- MEDIUM priority: low intra-state CV ---
MEDIUM_NOISE = {
    'Median Age Estimate':       (0.04, 15.0, 45.0, 'urban_mod'),
    '% Population Under 30':     (0.04, 25.0, 75.0, 'urban_mod'),
    'Fertility Rate Est':        (0.05, 1.0,  5.5,  'urban_mod'),
    'Gini Proxy':                (0.06, 0.15, 0.65, 'urban_mod'),
    'Adult Literacy Rate Pct':   (0.04, 0.0, 100.0, 'urban_mod'),
    'Male Literacy Rate Pct':    (0.04, 0.0, 100.0, 'urban_mod'),
    'Female Literacy Rate Pct':  (0.05, 0.0, 100.0, 'urban_mod'),
    'Primary Enrollment Pct':    (0.04, 0.0, 100.0, 'urban_mod'),
    'Secondary Enrollment Pct':  (0.05, 0.0, 100.0, 'urban_mod'),
    'Gender Parity Index':       (0.04, 0.3, 1.3,   'urban_mod'),
    'Youth Unemployment Rate Pct': (0.05, 0.0, 80.0, 'urban_mod'),
    'Out of School Children Pct':  (0.05, 0.0, 90.0, 'urban_mod'),
    'Housing Affordability':     (0.06, 0.0, 10.0,  'urban_mod'),
}

# --- LOW priority: columns that vary somewhat but could use more texture ---
LOW_NOISE = {
    'Poverty Rate Pct':          (0.03, 0.0, 100.0, 'urban_mod'),
    'Unemployment Rate Pct':     (0.03, 0.0, 80.0,  'urban_mod'),
    'Access Electricity Pct':    (0.03, 0.0, 100.0, 'urban_mod'),
    'Access Water Pct':          (0.03, 0.0, 100.0, 'urban_mod'),
    'Access Healthcare Pct':     (0.03, 0.0, 100.0, 'urban_mod'),
    'Road Quality Index':        (0.04, 1.0, 10.0,  'urban_mod'),
    'Market Access Index':       (0.04, 1.0, 10.0,  'urban_mod'),
    'Mobile Phone Penetration Pct': (0.02, 0.0, 100.0, 'absolute'),
    'Internet Access Pct':       (0.03, 0.0, 100.0, 'urban_mod'),
    'GDP Per Capita Est':        (0.08, 100, 200000, 'proportional'),
    'Biological Enhancement Pct': (0.05, 0.0, 100.0, 'urban_mod'),
    'Extraction Intensity':      (0.06, 0, 5, 'absolute'),
    'Land Formalization Pct':    (0.04, 0.0, 100.0, 'urban_mod'),
    'Trad Authority Index':      (0.08, 0, 5, 'absolute'),
    'Tijaniyya Presence':        (0.08, 0, 3, 'absolute'),
    'Qadiriyya Presence':        (0.08, 0, 3, 'absolute'),
    'Pentecostal Growth':        (0.08, 0, 3, 'absolute'),
    'Traditionalist Practice':   (0.08, 0, 3, 'absolute'),
    'Num Secondary Schools':     (0.10, 0, 200, 'proportional'),
    'Population Density per km2': (0.05, 1, 50000, 'proportional'),
}

# Columns whose noise must be coordinated (sum constraints)
# After adding noise to individual religious %s, re-normalize
RELIGION_COLS = ['% Muslim', '% Christian', '% Traditionalist']
LIVELIHOOD_COLS = ['Pct Livelihood Agriculture', 'Pct Livelihood Manufacturing',
                   'Pct Livelihood Extraction', 'Pct Livelihood Services',
                   'Pct Livelihood Informal']

# Ethnic columns that should get small noise where nonzero
ETHNIC_PREFIX = '% '
ETHNIC_EXCLUDE = {'% Muslim', '% Christian', '% Traditionalist',
                  '% Population Under 30'}

# ============================================================================
# HELPERS
# ============================================================================

def get_col_idx(headers, name):
    """Get 0-based index for a column name."""
    try:
        return headers.index(name)
    except ValueError:
        return None


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def compute_urban_modifier(urban_pct, pop_density):
    """LGAs that are more urban or more dense get slightly different noise
    than rural ones. Returns a modifier in [0.5, 1.5]."""
    # Urban areas: more variation in economic/cultural metrics
    # Rural areas: more variation in traditional/agricultural metrics
    u = min(100, max(0, safe_float(urban_pct, 30)))
    # Map 0-100 urban to 0.6-1.4 modifier
    return 0.6 + (u / 100.0) * 0.8


def apply_noise_to_column(data, rows, col_idx, urban_col_idx, density_col_idx,
                           std_frac, min_val, max_val, noise_type):
    """Apply Gaussian noise to a single column across all data rows."""
    col_range = max_val - min_val
    if col_range <= 0:
        return

    for row_idx in rows:
        row = data[row_idx]
        val = safe_float(row[col_idx])

        if noise_type == 'proportional':
            # Scale noise by current value
            std = max(abs(val) * std_frac, col_range * 0.005)
        elif noise_type == 'urban_mod':
            urban_pct = safe_float(row[urban_col_idx], 30) if urban_col_idx is not None else 30
            density = safe_float(row[density_col_idx], 100) if density_col_idx is not None else 100
            mod = compute_urban_modifier(urban_pct, density)
            std = col_range * std_frac * mod
        else:  # absolute
            std = col_range * std_frac

        noise = rng.normal(0, std)
        new_val = val + noise

        # Clamp
        new_val = max(min_val, min(max_val, new_val))

        # For integer columns, round
        if isinstance(min_val, int) and isinstance(max_val, int):
            new_val = int(round(new_val))
        else:
            new_val = round(new_val, 2)

        row[col_idx] = new_val


def apply_group_noise_and_normalize(data, rows, col_indices, target_sum,
                                     std_frac, urban_col_idx, density_col_idx):
    """Apply noise to a group of columns that must sum to a target, then renormalize."""
    for row_idx in rows:
        row = data[row_idx]
        vals = [safe_float(row[ci]) for ci in col_indices]
        total = sum(vals)
        if total <= 0:
            continue

        urban_pct = safe_float(row[urban_col_idx], 30) if urban_col_idx is not None else 30
        mod = compute_urban_modifier(urban_pct, 0)

        # Add noise to each
        noisy = []
        for v in vals:
            if v > 0.1:  # only add noise to nonzero components
                std = max(v * std_frac * mod, 0.3)
                n = v + rng.normal(0, std)
                noisy.append(max(0.1, n))
            else:
                noisy.append(max(0, v + rng.normal(0, 0.05)))

        # Renormalize to target_sum
        noisy_total = sum(noisy)
        if noisy_total > 0:
            factor = target_sum / noisy_total
            noisy = [n * factor for n in noisy]

        for ci, nv in zip(col_indices, noisy):
            row[ci] = round(max(0, nv), 2)


def apply_ethnic_noise(data, rows, headers, urban_col_idx, density_col_idx):
    """Add small noise to ethnic percentage columns, then renormalize to ~100."""
    # Find all ethnic columns (start with '% ', exclude religion/demographic)
    ethnic_indices = []
    for i, h in enumerate(headers):
        if h and h.startswith('% ') and h not in ETHNIC_EXCLUDE:
            ethnic_indices.append(i)

    for row_idx in rows:
        row = data[row_idx]
        urban_pct = safe_float(row[urban_col_idx], 30) if urban_col_idx is not None else 30
        mod = compute_urban_modifier(urban_pct, 0)

        vals = [safe_float(row[ci]) for ci in ethnic_indices]
        total = sum(vals)
        if total <= 0:
            continue

        noisy = []
        for v in vals:
            if v > 1.0:
                # Larger groups get proportionally smaller noise
                std = max(v * 0.03 * mod, 0.1)
                n = v + rng.normal(0, std)
                noisy.append(max(0, n))
            elif v > 0.01:
                # Small groups: tiny noise
                n = v + rng.normal(0, 0.05)
                noisy.append(max(0, n))
            else:
                noisy.append(0)

        # Renormalize
        noisy_total = sum(noisy)
        if noisy_total > 0:
            factor = 100.0 / noisy_total
            noisy = [n * factor for n in noisy]

        for ci, nv in zip(ethnic_indices, noisy):
            row[ci] = round(max(0, nv), 2)


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 60)
    print("Adding Gaussian noise to nigeria_lga_polsim_2058.xlsx")
    print("=" * 60)

    wb = load_workbook(INPUT_FILE)
    ws = wb['LGA_DATA']

    # Read headers (row 2)
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    num_cols = len(headers)

    # Read all data into memory
    data = []
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, num_cols + 1)]
        data.append(row)

    num_rows = len(data)
    all_rows = list(range(num_rows))
    print(f"Loaded {num_rows} LGAs, {num_cols} columns")

    # Key column indices
    urban_col = get_col_idx(headers, 'Urban Pct')
    density_col = get_col_idx(headers, 'Population Density per km2')
    state_col = get_col_idx(headers, 'State')

    # Combine all noise profiles
    all_profiles = {}
    all_profiles.update(HIGH_NOISE)
    all_profiles.update(MEDIUM_NOISE)
    all_profiles.update(LOW_NOISE)

    # --- Pass 1: Independent columns ---
    applied = 0
    for col_name, (std_frac, min_val, max_val, noise_type) in all_profiles.items():
        col_idx = get_col_idx(headers, col_name)
        if col_idx is None:
            print(f"  WARNING: Column '{col_name}' not found, skipping")
            continue
        apply_noise_to_column(data, all_rows, col_idx, urban_col, density_col,
                               std_frac, min_val, max_val, noise_type)
        applied += 1

    print(f"Applied noise to {applied} independent columns")

    # --- Pass 2: Religion group (must sum to ~100) ---
    relig_indices = [get_col_idx(headers, c) for c in RELIGION_COLS]
    if all(i is not None for i in relig_indices):
        apply_group_noise_and_normalize(data, all_rows, relig_indices, 100.0,
                                         0.03, urban_col, density_col)
        print("Applied noise to religion columns (renormalized)")

    # --- Pass 3: Livelihood group (must sum to ~100) ---
    liv_indices = [get_col_idx(headers, c) for c in LIVELIHOOD_COLS]
    if all(i is not None for i in liv_indices):
        apply_group_noise_and_normalize(data, all_rows, liv_indices, 100.0,
                                         0.04, urban_col, density_col)
        print("Applied noise to livelihood columns (renormalized)")

    # --- Pass 4: Ethnic columns (must sum to ~100) ---
    apply_ethnic_noise(data, all_rows, headers, urban_col, density_col)
    print("Applied noise to ethnic columns (renormalized)")

    # --- Pass 5: Consistency fixes ---
    median_age_col = get_col_idx(headers, 'Median Age Estimate')
    under30_col = get_col_idx(headers, '% Population Under 30')
    male_lit_col = get_col_idx(headers, 'Male Literacy Rate Pct')
    female_lit_col = get_col_idx(headers, 'Female Literacy Rate Pct')
    adult_lit_col = get_col_idx(headers, 'Adult Literacy Rate Pct')
    fert_col = get_col_idx(headers, 'Fertility Rate Est')
    poverty_col = get_col_idx(headers, 'Poverty Rate Pct')
    unemp_col = get_col_idx(headers, 'Unemployment Rate Pct')
    youth_unemp_col = get_col_idx(headers, 'Youth Unemployment Rate Pct')

    for row in data:
        # Median age <-> % under 30: higher age = lower under-30
        if median_age_col is not None and under30_col is not None:
            ma = safe_float(row[median_age_col])
            u30 = safe_float(row[under30_col])
            # Enforce inverse relationship: rough formula
            expected_u30 = max(25, min(75, 85 - ma * 1.2))
            # Blend: 70% noise result, 30% consistency enforcement
            row[under30_col] = round(u30 * 0.7 + expected_u30 * 0.3, 1)

        # Adult literacy should be between male and female (roughly)
        if all(c is not None for c in [male_lit_col, female_lit_col, adult_lit_col]):
            ml = safe_float(row[male_lit_col])
            fl = safe_float(row[female_lit_col])
            # Adult = weighted avg (assume ~50/50)
            expected_al = (ml + fl) / 2
            al = safe_float(row[adult_lit_col])
            row[adult_lit_col] = round(al * 0.5 + expected_al * 0.5, 1)
            # Ensure male >= female (in most LGAs)
            if ml < fl - 2:
                row[male_lit_col] = round(fl + rng.uniform(0, 3), 1)

        # Youth unemployment >= general unemployment
        if unemp_col is not None and youth_unemp_col is not None:
            u = safe_float(row[unemp_col])
            yu = safe_float(row[youth_unemp_col])
            if yu < u:
                row[youth_unemp_col] = round(u + rng.uniform(1, 5), 1)

        # Fertility clamping
        if fert_col is not None:
            row[fert_col] = round(max(1.0, min(5.5, safe_float(row[fert_col]))), 2)

        # Percentage clamping for all pct columns
        for ci in range(num_cols):
            h = headers[ci]
            if h and ('Pct' in h or h.startswith('%')):
                v = safe_float(row[ci])
                row[ci] = round(max(0, min(100, v)), 2)

    print("Applied consistency fixes")

    # --- Write back to sheet ---
    print("Writing data back to Excel...")
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            ws.cell(row=r_idx + 3, column=c_idx + 1, value=val)

    wb.save(OUTPUT_FILE)
    print(f"Saved to {OUTPUT_FILE}")

    # --- Verification ---
    print("\n=== VERIFICATION ===")
    # Check a few columns for intra-state variation
    state_idx = get_col_idx(headers, 'State')
    check_cols = ['Chinese Economic Presence', 'Mandarin Presence', 'Arabic Prestige',
                  'BIC Effectiveness', 'Al-Shahid Influence', 'English Prestige',
                  'Median Age Estimate', 'Fertility Rate Est', 'Gini Proxy',
                  'Adult Literacy Rate Pct', 'GDP Per Capita Est']
    for col_name in check_cols:
        ci = get_col_idx(headers, col_name)
        if ci is None:
            continue
        # Compute mean intra-state CV
        from collections import defaultdict
        state_vals = defaultdict(list)
        for row in data:
            s = row[state_idx]
            v = safe_float(row[ci])
            state_vals[s].append(v)
        cvs = []
        identical_states = 0
        for s, vals in state_vals.items():
            if len(vals) < 2:
                continue
            mean = np.mean(vals)
            std = np.std(vals)
            if mean > 0:
                cvs.append(std / mean)
            if std < 0.001:
                identical_states += 1
        mean_cv = np.mean(cvs) if cvs else 0
        print(f"  {col_name:35s} mean CV={mean_cv:.4f}  identical_states={identical_states}")

    # Check sums
    relig_ok = 0
    relig_bad = 0
    for row in data:
        rs = sum(safe_float(row[ri]) for ri in relig_indices if ri is not None)
        if 95 <= rs <= 105:
            relig_ok += 1
        else:
            relig_bad += 1
    print(f"\n  Religion sums: {relig_ok} OK, {relig_bad} bad (out of {num_rows})")

    ethnic_indices = [i for i, h in enumerate(headers)
                      if h and h.startswith('% ') and h not in ETHNIC_EXCLUDE]
    eth_ok = 0
    eth_bad = 0
    for row in data:
        es = sum(safe_float(row[ei]) for ei in ethnic_indices)
        if 95 <= es <= 105:
            eth_ok += 1
        else:
            eth_bad += 1
    print(f"  Ethnic sums: {eth_ok} OK, {eth_bad} bad (out of {num_rows})")

    liv_ok = 0
    liv_bad = 0
    for row in data:
        ls = sum(safe_float(row[li]) for li in liv_indices if li is not None)
        if 95 <= ls <= 105:
            liv_ok += 1
        else:
            liv_bad += 1
    print(f"  Livelihood sums: {liv_ok} OK, {liv_bad} bad (out of {num_rows})")

    print("\nDone.")


if __name__ == '__main__':
    main()
