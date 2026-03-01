#!/usr/bin/env python3
"""
Fix script for nigeria_lga_polsim_2058.xlsx
Addresses all verified problems in the data.
Reads the spreadsheet, applies corrections, writes back.
"""
import math
import numpy as np
from openpyxl import load_workbook
from collections import defaultdict

INPUT_FILE = "nigeria_lga_polsim_2058.xlsx"
OUTPUT_FILE = "nigeria_lga_polsim_2058.xlsx"  # overwrite in place
SEED = 9999
rng = np.random.default_rng(SEED)

NORTHERN_STATES = {
    'Adamawa', 'Bauchi', 'Benue', 'Borno', 'FCT', 'Gombe', 'Jigawa',
    'Kaduna', 'Kano', 'Katsina', 'Kebbi', 'Kogi', 'Kwara', 'Nasarawa',
    'Niger', 'Plateau', 'Sokoto', 'Taraba', 'Yobe', 'Zamfara'
}
SOUTHERN_STATES = {
    'Abia', 'Akwa Ibom', 'Anambra', 'Bayelsa', 'Cross River', 'Delta',
    'Ebonyi', 'Edo', 'Ekiti', 'Enugu', 'Imo', 'Lagos', 'Ogun', 'Ondo',
    'Osun', 'Oyo', 'Rivers'
}

# Conflict-affected areas for calibration
AL_SHAHID_STATES = {'Kano'}
SEVERE_CONFLICT_STATES = {'Borno', 'Yobe', 'Adamawa', 'Zamfara'}
MODERATE_CONFLICT_STATES = {'Katsina', 'Kaduna', 'Niger', 'Sokoto', 'Kebbi'}

# Terrain types that should have poor roads
POOR_ROAD_TERRAINS = {'mangrove', 'swamp', 'riverine', 'creekland', 'wetland',
                       'floodplain', 'island', 'coastal', 'deltaic'}

# Authorized refinery locations only
AUTHORIZED_REFINERIES = {
    ('Rivers', 'Bonny'), ('Rivers', 'Degema'),
    ('Kogi', 'Lokoja'), ('Kogi', 'Ajaokuta'),
    ('Borno', 'Maiduguri'), ('Borno', 'Jere'),
}

# Lagos LGAs that should have higher Pada concentrations
LAGOS_WEALTHY_LGAS = {'Eti-Osa', 'Ibeju-Lekki', 'Lagos Island', 'Apapa', 'Ikeja'}
LAGOS_WORKING_CLASS_LGAS = {'Agege', 'Mushin', 'Ajeromi-Ifelodun', 'Oshodi-Isolo',
                             'Alimosho', 'Kosofe', 'Shomolu', 'Ifako-Ijaiye'}


def sf(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def main():
    print("=" * 60)
    print("FIXING nigeria_lga_polsim_2058.xlsx")
    print("=" * 60)

    wb = load_workbook(INPUT_FILE)
    ws = wb['LGA_DATA']

    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    num_cols = len(headers)

    def ci(name):
        try:
            return headers.index(name)
        except ValueError:
            print(f"  WARNING: Column '{name}' not found")
            return None

    # Read all data into memory
    data = []
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, num_cols + 1)]
        if row[0] is not None:
            data.append(row)

    num_rows = len(data)
    print(f"Loaded {num_rows} LGAs, {num_cols} columns")

    # Key column indices
    STATE = ci('State')
    LGA = ci('LGA Name')
    TERRAIN = ci('Terrain Type')
    POP = ci('Estimated Population')
    GDP = ci('GDP Per Capita Est')
    POVERTY = ci('Poverty Rate Pct')
    LITERACY = ci('Adult Literacy Rate Pct')
    MALE_LIT = ci('Male Literacy Rate Pct')
    FEMALE_LIT = ci('Female Literacy Rate Pct')
    FERTILITY = ci('Fertility Rate Est')
    SHAHID = ci('Al-Shahid Influence')
    URBAN = ci('Urban Pct')
    ROAD = ci('Road Quality Index')
    COBALT = ci('Cobalt Extraction Active')
    REFINERY = ci('Refinery Present')
    REFINERY_ZONE = ci('Refinery Zone')
    CONFLICT = ci('Conflict History')
    ALMAJIRI = ci('Almajiri Index')
    BIO = ci('Biological Enhancement Pct')
    PADA = ci('% Pada')
    NAIJIN = ci('% Naijin')
    PRIMARY_ENROLL = ci('Primary Enrollment Pct')
    SECONDARY_ENROLL = ci('Secondary Enrollment Pct')
    GENDER_PARITY = ci('Gender Parity Index')
    OUT_OF_SCHOOL = ci('Out of School Children Pct')
    DENSITY = ci('Population Density per km2')
    MEDIAN_AGE = ci('Median Age Estimate')
    UNDER_30 = ci('% Population Under 30')
    EXTRACTION_INTENSITY = ci('Extraction Intensity')
    OIL_ACTIVE = ci('Oil Extraction Active')
    OTHER_MINING = ci('Other Mining Active')

    # ================================================================
    # FIX 1: GDP PER CAPITA - Scale down to hit $26-28k weighted avg
    # ================================================================
    print("\n--- FIX 1: GDP Per Capita ---")
    # Current weighted avg is $41,953, target is ~$27,000
    # Compute current weighted average
    total_gdp_pop = sum(sf(r[GDP]) * sf(r[POP]) for r in data)
    total_pop = sum(sf(r[POP]) for r in data)
    current_avg = total_gdp_pop / total_pop if total_pop > 0 else 1
    target_avg = 27000.0
    # Simple scaling factor
    scale = target_avg / current_avg
    print(f"  Current weighted avg: ${current_avg:,.0f}")
    print(f"  Scale factor: {scale:.4f}")

    for r in data:
        old_gdp = sf(r[GDP])
        # Apply scale, but don't let any LGA go below $500
        r[GDP] = round(max(500, old_gdp * scale), 0)

    # Verify
    new_avg = sum(sf(r[GDP]) * sf(r[POP]) for r in data) / total_pop
    print(f"  New weighted avg: ${new_avg:,.0f}")

    # ================================================================
    # FIX 2: NORTHERN POVERTY - Reduce to 30-45% range
    # ================================================================
    print("\n--- FIX 2: Northern Poverty ---")
    # Current north mean: 61.1%, target: 30-45%
    # Strategy: non-conflict northern LGAs get substantial reduction
    # Al-Shahid/conflict LGAs get less reduction but still some
    for r in data:
        state = r[STATE]
        if state not in NORTHERN_STATES:
            continue

        old_pov = sf(r[POVERTY])
        conflict = r[CONFLICT] if CONFLICT else 'SPARED'
        lga_name = r[LGA]

        if conflict in ('AL_SHAHID_CONTROL',):
            # Al-Shahid zones: still high poverty but reduced from 2026
            # Target: 50-65% (was originally very high)
            target = max(45, min(65, old_pov * 0.85 + rng.normal(0, 3)))
        elif conflict in ('SEVERE', 'OCCUPATION'):
            # Severe conflict: limited development dividend
            # Target: 40-55%
            target = max(35, min(55, old_pov * 0.75 + rng.normal(0, 3)))
        elif conflict == 'LOW_IMPACT':
            # Some conflict impact but development still reached
            # Target: 30-45%
            target = max(25, min(45, old_pov * 0.60 + rng.normal(0, 3)))
        else:
            # SPARED - full development dividend
            # Target: 25-40% (industrialization, rail, cobalt revenue)
            urban = sf(r[URBAN])
            if urban > 60:
                # Urban northern LGAs - most benefit
                target = max(18, min(35, old_pov * 0.45 + rng.normal(0, 3)))
            elif urban > 30:
                target = max(22, min(40, old_pov * 0.55 + rng.normal(0, 3)))
            else:
                # Rural northern - still significant improvement
                target = max(28, min(45, old_pov * 0.60 + rng.normal(0, 3)))

        r[POVERTY] = round(max(3, min(85, target)), 1)

    north_pov = [sf(r[POVERTY]) for r in data if r[STATE] in NORTHERN_STATES]
    print(f"  New north mean poverty: {np.mean(north_pov):.1f}%")

    # ================================================================
    # FIX 3: NORTHERN LITERACY - Increase to 65-75% range
    # ================================================================
    print("\n--- FIX 3: Northern Literacy ---")
    # Current north mean: 50.3%, target: 65-75%
    for r in data:
        state = r[STATE]
        if state not in NORTHERN_STATES:
            continue

        old_lit = sf(r[LITERACY])
        old_male = sf(r[MALE_LIT])
        old_female = sf(r[FEMALE_LIT])
        conflict = r[CONFLICT] if CONFLICT else 'SPARED'
        urban = sf(r[URBAN])

        if conflict in ('AL_SHAHID_CONTROL',):
            # Al-Shahid zones: education regressed
            # Target: 40-55%
            target_lit = max(35, min(55, old_lit * 1.0 + rng.normal(0, 2)))
            female_gap = -15  # women's education suffered
        elif conflict in ('SEVERE', 'OCCUPATION'):
            # Severe conflict: limited education gains
            # Target: 55-65%
            target_lit = max(50, min(65, old_lit * 1.15 + rng.normal(0, 2)))
            female_gap = -10
        elif conflict == 'LOW_IMPACT':
            # Some impact but education still advanced
            # Target: 60-75%
            target_lit = max(58, min(78, old_lit * 1.30 + rng.normal(0, 2)))
            female_gap = -7
        else:
            # SPARED - full education dividend
            if urban > 60:
                target_lit = max(72, min(90, old_lit * 1.50 + rng.normal(0, 2)))
                female_gap = -4
            elif urban > 30:
                target_lit = max(65, min(82, old_lit * 1.40 + rng.normal(0, 2)))
                female_gap = -6
            else:
                target_lit = max(60, min(78, old_lit * 1.35 + rng.normal(0, 2)))
                female_gap = -8

        # Set adult literacy
        r[LITERACY] = round(max(30, min(98, target_lit)), 1)

        # Male literacy should be somewhat higher than adult
        new_male = r[LITERACY] + abs(female_gap) / 2 + rng.normal(0, 1)
        r[MALE_LIT] = round(max(r[LITERACY], min(99, new_male)), 1)

        # Female literacy: adult - gap
        new_female = r[LITERACY] + female_gap + rng.normal(0, 1)
        r[FEMALE_LIT] = round(max(25, min(95, new_female)), 1)

        # Also update enrollment rates for consistency
        if PRIMARY_ENROLL:
            old_pe = sf(r[PRIMARY_ENROLL])
            if conflict not in ('AL_SHAHID_CONTROL',):
                r[PRIMARY_ENROLL] = round(min(99, max(old_pe, r[LITERACY] * 1.05 + rng.normal(0, 2))), 1)
        if SECONDARY_ENROLL:
            old_se = sf(r[SECONDARY_ENROLL])
            if conflict not in ('AL_SHAHID_CONTROL',):
                r[SECONDARY_ENROLL] = round(min(95, max(old_se, r[LITERACY] * 0.75 + rng.normal(0, 3))), 1)
        if OUT_OF_SCHOOL:
            # Out of school inversely related to literacy
            r[OUT_OF_SCHOOL] = round(max(3, min(60, 100 - r[LITERACY] * 0.9 + rng.normal(0, 3))), 1)
        if GENDER_PARITY:
            gp = sf(r[FEMALE_LIT]) / max(1, sf(r[MALE_LIT]))
            r[GENDER_PARITY] = round(max(0.4, min(1.2, gp + rng.normal(0, 0.02))), 2)

    north_lit = [sf(r[LITERACY]) for r in data if r[STATE] in NORTHERN_STATES]
    print(f"  New north mean literacy: {np.mean(north_lit):.1f}%")

    # ================================================================
    # FIX 4: TOTAL POPULATION - Scale up to 280-320M
    # ================================================================
    print("\n--- FIX 4: Total Population ---")
    current_total = sum(sf(r[POP]) for r in data)
    target_total = 295e6  # middle of 280-320M range
    # North should have grown faster (sustained ~3.0 fertility)
    # Scale north more than south
    north_pop = sum(sf(r[POP]) for r in data if r[STATE] in NORTHERN_STATES)
    south_pop = sum(sf(r[POP]) for r in data if r[STATE] in SOUTHERN_STATES)
    print(f"  Current total: {current_total/1e6:.1f}M (N:{north_pop/1e6:.1f}M, S:{south_pop/1e6:.1f}M)")

    # We need to add ~35M. North should get ~60% of the increase
    needed = target_total - current_total
    north_share = 0.60
    north_scale = 1.0 + (needed * north_share) / north_pop
    south_scale = 1.0 + (needed * (1 - north_share)) / south_pop

    for r in data:
        if r[STATE] in NORTHERN_STATES:
            r[POP] = round(sf(r[POP]) * north_scale)
        else:
            r[POP] = round(sf(r[POP]) * south_scale)
        # Update density proportionally
        if DENSITY:
            r[DENSITY] = round(sf(r[DENSITY]) * (north_scale if r[STATE] in NORTHERN_STATES else south_scale), 1)

    new_total = sum(sf(r[POP]) for r in data)
    new_north = sum(sf(r[POP]) for r in data if r[STATE] in NORTHERN_STATES)
    new_south = sum(sf(r[POP]) for r in data if r[STATE] in SOUTHERN_STATES)
    print(f"  New total: {new_total/1e6:.1f}M (N:{new_north/1e6:.1f}M, S:{new_south/1e6:.1f}M)")

    # ================================================================
    # FIX 5: POVERTY FLOOR - No LGA at 0%
    # ================================================================
    print("\n--- FIX 5: Poverty Floor ---")
    zero_count_before = sum(1 for r in data if sf(r[POVERTY]) < 1)
    for r in data:
        pov = sf(r[POVERTY])
        state = r[STATE]
        urban = sf(r[URBAN])

        if pov < 2:
            # Wealthy urban LGAs: floor of 3-5%
            if state == 'Lagos' or urban > 80:
                floor = 3.0 + rng.uniform(0, 3)
            elif state in SOUTHERN_STATES:
                floor = 4.0 + rng.uniform(0, 4)
            else:
                floor = 5.0 + rng.uniform(0, 5)
            r[POVERTY] = round(max(floor, pov), 1)

    zero_count_after = sum(1 for r in data if sf(r[POVERTY]) < 1)
    print(f"  LGAs < 1% poverty: {zero_count_before} -> {zero_count_after}")

    # ================================================================
    # FIX 6: FERTILITY FLOOR - Clamp at 1.3
    # ================================================================
    print("\n--- FIX 6: Fertility Floor ---")
    below_13_before = sum(1 for r in data if sf(r[FERTILITY]) < 1.3)
    for r in data:
        fert = sf(r[FERTILITY])
        if fert < 1.3:
            # Clamp to 1.3-1.5 with small noise
            r[FERTILITY] = round(1.3 + rng.uniform(0, 0.15), 2)

    below_13_after = sum(1 for r in data if sf(r[FERTILITY]) < 1.3)
    print(f"  LGAs below 1.3: {below_13_before} -> {below_13_after}")

    # ================================================================
    # FIX 7: AL-SHAHID INFLUENCE IN KANO
    # ================================================================
    print("\n--- FIX 7: Al-Shahid Influence in Kano ---")
    # Kano fell in 2040, al-shahid influence jumped to 4-5
    # 2052 peace talks reduced by 0.5 -> by 2058 should be 3.5-4.5
    for r in data:
        if r[STATE] != 'Kano':
            continue
        old_val = sf(r[SHAHID])
        # Target: 3.5-4.5 for most Kano LGAs
        urban = sf(r[URBAN])
        muslim_pct = sf(r[ci('% Muslim')]) if ci('% Muslim') else 80

        if muslim_pct > 80:
            # High Muslim areas: strongest influence
            new_val = 4.0 + rng.uniform(-0.3, 0.5)
        elif muslim_pct > 60:
            new_val = 3.7 + rng.uniform(-0.3, 0.3)
        else:
            # Minority areas: somewhat lower
            new_val = 3.2 + rng.uniform(-0.3, 0.3)

        r[SHAHID] = round(max(2.5, min(5.0, new_val)), 2)

    kano_shahid = [sf(r[SHAHID]) for r in data if r[STATE] == 'Kano']
    print(f"  Kano al-shahid mean: {np.mean(kano_shahid):.2f}, range: {min(kano_shahid):.2f}-{max(kano_shahid):.2f}")

    # ================================================================
    # FIX 8: LAGOS URBANIZATION
    # ================================================================
    print("\n--- FIX 8: Lagos Urbanization ---")
    for r in data:
        if r[STATE] != 'Lagos':
            continue
        old_urban = sf(r[URBAN])
        lga_name = r[LGA]

        # By 2058, every Lagos LGA should be 85%+ urban
        if old_urban < 85:
            if 'Ibeju' in str(lga_name):
                # New Makoko site: should be 90%+ by now
                r[URBAN] = round(90 + rng.uniform(0, 5), 1)
            elif 'Badagry' in str(lga_name):
                r[URBAN] = round(87 + rng.uniform(0, 5), 1)
            elif 'Epe' in str(lga_name):
                r[URBAN] = round(85 + rng.uniform(0, 5), 1)
            elif 'Ikorodu' in str(lga_name):
                r[URBAN] = round(90 + rng.uniform(0, 5), 1)
            else:
                r[URBAN] = round(max(85, old_urban + rng.uniform(0, 5)), 1)
        # Cap at 100
        r[URBAN] = min(100, sf(r[URBAN]))

    lagos_urban = [sf(r[URBAN]) for r in data if r[STATE] == 'Lagos']
    print(f"  Lagos urban: min={min(lagos_urban):.1f}%, mean={np.mean(lagos_urban):.1f}%")

    # ================================================================
    # FIX 9: SOUTHERN ROAD QUALITY - Add variation
    # ================================================================
    print("\n--- FIX 9: Southern Road Quality ---")
    for r in data:
        if r[STATE] not in SOUTHERN_STATES:
            continue

        old_road = sf(r[ROAD])
        terrain = str(r[TERRAIN] or '').lower()
        urban = sf(r[URBAN])
        state = r[STATE]

        # Check for poor-road terrains
        is_difficult = any(t in terrain for t in POOR_ROAD_TERRAINS)

        if is_difficult and urban < 40:
            # Remote riverine/swamp/mangrove areas
            new_road = 4.5 + rng.uniform(-0.5, 1.5)
        elif is_difficult:
            # Urban but difficult terrain
            new_road = 6.0 + rng.uniform(-0.5, 1.5)
        elif state in ('Bayelsa',) and urban < 50:
            # Bayelsa creekland
            new_road = 5.0 + rng.uniform(-0.5, 1.0)
        elif urban > 80:
            # Major urban: excellent roads
            new_road = 8.5 + rng.uniform(0, 1.5)
        elif urban > 50:
            # Semi-urban: good roads
            new_road = 7.5 + rng.uniform(-0.5, 1.5)
        elif urban > 25:
            # Semi-rural: decent
            new_road = 6.5 + rng.uniform(-0.5, 1.5)
        else:
            # Rural: moderate
            new_road = 5.5 + rng.uniform(-0.5, 1.5)

        r[ROAD] = round(max(1.0, min(10.0, new_road)), 1)

    south_roads = [sf(r[ROAD]) for r in data if r[STATE] in SOUTHERN_STATES]
    print(f"  South road: min={min(south_roads):.1f}, max={max(south_roads):.1f}, mean={np.mean(south_roads):.1f}, std={np.std(south_roads):.2f}")

    # ================================================================
    # FIX 10: REMOVE COBALT FROM KATSINA
    # ================================================================
    print("\n--- FIX 10: Cobalt in Katsina ---")
    removed = 0
    for r in data:
        if r[STATE] == 'Katsina' and sf(r[COBALT]) > 0:
            r[COBALT] = 0
            # Also reduce extraction intensity if it was cobalt-driven
            if EXTRACTION_INTENSITY:
                ei = sf(r[EXTRACTION_INTENSITY])
                if sf(r[OIL_ACTIVE]) == 0 and sf(r[OTHER_MINING]) == 0:
                    r[EXTRACTION_INTENSITY] = 0
                elif ei > 2:
                    r[EXTRACTION_INTENSITY] = max(1, ei - 2)
            removed += 1
            print(f"  Removed cobalt from {r[STATE]}-{r[LGA]}")
    print(f"  Total removed: {removed}")

    # ================================================================
    # FIX 11: REMOVE UNAUTHORIZED REFINERIES
    # ================================================================
    print("\n--- FIX 11: Unauthorized Refineries ---")
    removed = 0
    for r in data:
        if sf(r[REFINERY]) > 0:
            key = (r[STATE], r[LGA])
            if key not in AUTHORIZED_REFINERIES:
                r[REFINERY] = 0
                if REFINERY_ZONE:
                    r[REFINERY_ZONE] = 0
                # Reduce extraction intensity if refinery was the driver
                if EXTRACTION_INTENSITY:
                    ei = sf(r[EXTRACTION_INTENSITY])
                    if sf(r[OIL_ACTIVE]) == 0 and sf(r[COBALT]) == 0 and sf(r[OTHER_MINING]) == 0:
                        r[EXTRACTION_INTENSITY] = max(0, ei - 2)
                removed += 1
                print(f"  Removed refinery from {r[STATE]}-{r[LGA]}")
    print(f"  Total removed: {removed}")

    # ================================================================
    # FIX 12: ADD MODERATE CONFLICT CATEGORY
    # ================================================================
    print("\n--- FIX 12: MODERATE Conflict Category ---")
    # Some LGAs between LOW_IMPACT and SEVERE should be MODERATE
    # Candidates: LGAs in conflict-adjacent states that had some fighting
    # but weren't fully devastated
    moderate_count = 0
    for r in data:
        state = r[STATE]
        conflict = r[CONFLICT]

        # Reassign some LOW_IMPACT to MODERATE in heavily affected states
        if conflict == 'LOW_IMPACT' and state in MODERATE_CONFLICT_STATES:
            r[CONFLICT] = 'MODERATE'
            moderate_count += 1
        # Also promote some SEVERE in less-affected states to MODERATE
        elif conflict == 'SEVERE' and state in ('Plateau', 'Nasarawa', 'Taraba', 'Gombe'):
            # These states had conflict but not the worst of it
            if rng.random() < 0.5:
                r[CONFLICT] = 'MODERATE'
                moderate_count += 1
        # And some SPARED in states that had minor spillover
        elif conflict == 'SPARED' and state in MODERATE_CONFLICT_STATES:
            if rng.random() < 0.3:
                r[CONFLICT] = 'LOW_IMPACT'
                # A few of these become MODERATE
                if rng.random() < 0.3:
                    r[CONFLICT] = 'MODERATE'
                    moderate_count += 1

    # Count categories
    cats = defaultdict(int)
    for r in data:
        cats[r[CONFLICT]] += 1
    print(f"  Added {moderate_count} MODERATE entries")
    print(f"  Categories: {dict(cats)}")

    # ================================================================
    # FIX 13: ALMAJIRI IN SOUTH
    # ================================================================
    print("\n--- FIX 13: Almajiri in South ---")
    # Most southern LGAs should be 0. Only ~20-30 urban southern LGAs at 1-2
    south_almajiri_before = sum(1 for r in data if r[STATE] in SOUTHERN_STATES and sf(r[ALMAJIRI]) > 0)

    # Major southern urban centers that might have some almajiri presence
    SOUTH_ALMAJIRI_CITIES = {
        'Lagos': ['Agege', 'Mushin', 'Oshodi-Isolo', 'Alimosho', 'Ifako-Ijaiye',
                   'Kosofe', 'Ikeja', 'Surulere', 'Ajeromi-Ifelodun', 'Lagos Mainland'],
        'Oyo': ['Ibadan North', 'Ibadan North-East', 'Ibadan South-West'],
        'Ogun': ['Abeokuta South', 'Abeokuta North'],
        'Edo': ['Oredo'],
        'Delta': ['Warri South'],
        'Rivers': ['Port Harcourt', 'Obio/Akpor'],
        'Anambra': ['Onitsha North', 'Onitsha South'],
        'Enugu': ['Enugu North', 'Enugu South'],
        'Abia': ['Aba North', 'Aba South'],
    }

    for r in data:
        if r[STATE] not in SOUTHERN_STATES:
            continue

        state = r[STATE]
        lga_name = str(r[LGA])

        # Check if this is one of the ~20-30 urban southern LGAs with almajiri
        is_almajiri_city = False
        if state in SOUTH_ALMAJIRI_CITIES:
            for city_lga in SOUTH_ALMAJIRI_CITIES[state]:
                if city_lga in lga_name or lga_name in city_lga:
                    is_almajiri_city = True
                    break

        if is_almajiri_city:
            # Level 1-2 for urban centers with northern migrant populations
            r[ALMAJIRI] = round(1 + rng.uniform(0, 1), 0)
        else:
            # All other southern LGAs: 0
            r[ALMAJIRI] = 0

    south_almajiri_after = sum(1 for r in data if r[STATE] in SOUTHERN_STATES and sf(r[ALMAJIRI]) > 0)
    print(f"  Southern LGAs with Almajiri > 0: {south_almajiri_before} -> {south_almajiri_after}")

    # Also fix northern almajiri - cap the level 5 epidemic
    # Most northern should be 3-5, not all at 5
    level5_before = sum(1 for r in data if sf(r[ALMAJIRI]) == 5)
    for r in data:
        if r[STATE] not in NORTHERN_STATES:
            continue
        alm = sf(r[ALMAJIRI])
        conflict = r[CONFLICT]
        urban = sf(r[URBAN])
        muslim = sf(r[ci('% Muslim')]) if ci('% Muslim') else 50

        if conflict in ('AL_SHAHID_CONTROL',):
            # Al-Shahid zones: highest almajiri
            r[ALMAJIRI] = round(min(5, 4 + rng.uniform(0, 1)), 0)
        elif muslim > 70 and conflict in ('SEVERE',):
            r[ALMAJIRI] = round(min(5, 3 + rng.uniform(0, 1.5)), 0)
        elif muslim > 70:
            if urban > 50:
                r[ALMAJIRI] = round(min(5, 2 + rng.uniform(0, 1.5)), 0)
            else:
                r[ALMAJIRI] = round(min(5, 3 + rng.uniform(0, 1.5)), 0)
        elif muslim > 40:
            r[ALMAJIRI] = round(min(4, 1 + rng.uniform(0, 1.5)), 0)
        elif muslim > 15:
            r[ALMAJIRI] = round(min(3, 0 + rng.uniform(0, 1)), 0)
        else:
            r[ALMAJIRI] = 0

    level5_after = sum(1 for r in data if sf(r[ALMAJIRI]) == 5)
    print(f"  Level 5 LGAs: {level5_before} -> {level5_after}")

    # ================================================================
    # FIX 14: BIOLOGICAL ENHANCEMENT - Scale to 42% target
    # ================================================================
    print("\n--- FIX 14: Biological Enhancement ---")
    bio_vals = [(sf(r[BIO]), sf(r[POP])) for r in data]
    current_bio_avg = sum(b * p for b, p in bio_vals) / sum(p for b, p in bio_vals)
    target_bio = 42.0
    bio_scale = target_bio / current_bio_avg if current_bio_avg > 0 else 1

    for r in data:
        old_bio = sf(r[BIO])
        r[BIO] = round(max(0, min(85, old_bio * bio_scale)), 1)

    new_bio_avg = sum(sf(r[BIO]) * sf(r[POP]) for r in data) / sum(sf(r[POP]) for r in data)
    print(f"  Bio enhancement weighted avg: {current_bio_avg:.1f}% -> {new_bio_avg:.1f}%")

    # ================================================================
    # FIX 15: PADA DISTRIBUTION IN LAGOS
    # ================================================================
    print("\n--- FIX 15: Pada Distribution in Lagos ---")
    # Wealthy LGAs (Eti-Osa, Ibeju-Lekki) should be higher
    # Working class LGAs (Agege, Mushin) should be lower
    # Need to redistribute ethnic % to maintain sum = 100

    # Get all ethnic column indices
    ethnic_start = ci('% Hausa')
    other_col = ci('% Other')
    ethnic_indices = list(range(ethnic_start, other_col + 1))  # all ethnic columns

    for r in data:
        if r[STATE] != 'Lagos':
            continue
        lga_name = str(r[LGA])
        old_pada = sf(r[PADA])

        if lga_name in LAGOS_WEALTHY_LGAS:
            # Wealthy areas: higher pada (12-15%)
            target_pada = 12.0 + rng.uniform(0, 3)
        elif lga_name in LAGOS_WORKING_CLASS_LGAS:
            # Working class: lower pada (6-8%)
            target_pada = 6.0 + rng.uniform(0, 2)
        else:
            # Middle: moderate (8-11%)
            target_pada = 8.5 + rng.uniform(0, 2.5)

        pada_diff = target_pada - old_pada
        r[PADA] = round(target_pada, 2)

        # Redistribute the difference across other ethnic groups proportionally
        other_total = sum(sf(r[i]) for i in ethnic_indices if i != PADA and i != NAIJIN)
        if other_total > 0 and abs(pada_diff) > 0.01:
            scale_factor = (other_total - pada_diff) / other_total
            for i in ethnic_indices:
                if i != PADA and i != NAIJIN:
                    r[i] = round(max(0, sf(r[i]) * scale_factor), 2)

    lagos_pada = [(r[LGA], sf(r[PADA])) for r in data if r[STATE] == 'Lagos']
    lagos_pada.sort(key=lambda x: -x[1])
    print("  Lagos Pada distribution (top/bottom):")
    for l, p in lagos_pada[:5]:
        print(f"    {l}: {p:.2f}%")
    print("    ...")
    for l, p in lagos_pada[-5:]:
        print(f"    {l}: {p:.2f}%")

    # ================================================================
    # FINAL: Write back to sheet
    # ================================================================
    print("\n--- Writing data back to Excel ---")
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            ws.cell(row=r_idx + 3, column=c_idx + 1, value=val)

    wb.save(OUTPUT_FILE)
    print(f"Saved to {OUTPUT_FILE}")

    # ================================================================
    # VERIFICATION
    # ================================================================
    print("\n" + "=" * 60)
    print("VERIFICATION SUMMARY")
    print("=" * 60)

    # GDP
    gdp_avg = sum(sf(r[GDP]) * sf(r[POP]) for r in data) / sum(sf(r[POP]) for r in data)
    print(f"  GDP weighted avg: ${gdp_avg:,.0f}  (target: $26-28k)")

    # Poverty
    n_pov = np.mean([sf(r[POVERTY]) for r in data if r[STATE] in NORTHERN_STATES])
    s_pov = np.mean([sf(r[POVERTY]) for r in data if r[STATE] in SOUTHERN_STATES])
    print(f"  North poverty: {n_pov:.1f}%  (target: 30-45%)")
    print(f"  South poverty: {s_pov:.1f}%")

    # Literacy
    n_lit = np.mean([sf(r[LITERACY]) for r in data if r[STATE] in NORTHERN_STATES])
    s_lit = np.mean([sf(r[LITERACY]) for r in data if r[STATE] in SOUTHERN_STATES])
    print(f"  North literacy: {n_lit:.1f}%  (target: 65-75%)")
    print(f"  South literacy: {s_lit:.1f}%")

    # Population
    total = sum(sf(r[POP]) for r in data) / 1e6
    print(f"  Total population: {total:.1f}M  (target: 280-320M)")

    # Poverty floor
    zero_pov = sum(1 for r in data if sf(r[POVERTY]) < 1)
    print(f"  LGAs < 1% poverty: {zero_pov}  (target: 0)")

    # Fertility floor
    below_13 = sum(1 for r in data if sf(r[FERTILITY]) < 1.3)
    print(f"  LGAs below 1.3 fertility: {below_13}  (target: 0)")

    # Kano al-shahid
    k_shahid = [sf(r[SHAHID]) for r in data if r[STATE] == 'Kano']
    print(f"  Kano al-shahid: mean={np.mean(k_shahid):.2f}  (target: 3.5-4.5)")

    # Lagos urbanization
    l_urban = [sf(r[URBAN]) for r in data if r[STATE] == 'Lagos']
    print(f"  Lagos urban: min={min(l_urban):.1f}%  (target: 85%+)")

    # South roads
    s_roads = [sf(r[ROAD]) for r in data if r[STATE] in SOUTHERN_STATES]
    print(f"  South road: min={min(s_roads):.1f}, mean={np.mean(s_roads):.1f}, std={np.std(s_roads):.2f}")

    # Cobalt
    cobalt_lgas = [(r[STATE], r[LGA]) for r in data if sf(r[COBALT]) > 0]
    print(f"  Cobalt LGAs: {cobalt_lgas}")

    # Refineries
    refinery_lgas = [(r[STATE], r[LGA]) for r in data if sf(r[REFINERY]) > 0]
    print(f"  Refinery LGAs: {refinery_lgas}")

    # Conflict categories
    cats = defaultdict(int)
    for r in data:
        cats[r[CONFLICT]] += 1
    print(f"  Conflict categories: {dict(cats)}")

    # Almajiri
    south_alm = sum(1 for r in data if r[STATE] in SOUTHERN_STATES and sf(r[ALMAJIRI]) > 0)
    level5 = sum(1 for r in data if sf(r[ALMAJIRI]) == 5)
    print(f"  Southern LGAs with Almajiri > 0: {south_alm}  (target: 20-30)")
    print(f"  Level 5 LGAs: {level5}")

    # Bio enhancement
    bio = sum(sf(r[BIO]) * sf(r[POP]) for r in data) / sum(sf(r[POP]) for r in data)
    print(f"  Bio enhancement weighted avg: {bio:.1f}%  (target: 42%)")

    print("\nDone.")


if __name__ == '__main__':
    main()
