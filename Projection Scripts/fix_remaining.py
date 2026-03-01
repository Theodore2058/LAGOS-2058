#!/usr/bin/env python3
"""
Fix remaining issues:
1. Northern literacy (60.6% -> 65-75%)
2. MODERATE conflict category (need to convert some SEVERE to MODERATE)
"""
import numpy as np
from openpyxl import load_workbook
from collections import defaultdict

INPUT_FILE = "nigeria_lga_polsim_2058.xlsx"
OUTPUT_FILE = "nigeria_lga_polsim_2058.xlsx"
SEED = 8888
rng = np.random.default_rng(SEED)

NORTHERN_STATES = {
    'Adamawa', 'Bauchi', 'Benue', 'Borno', 'FCT', 'Gombe', 'Jigawa',
    'Kaduna', 'Kano', 'Katsina', 'Kebbi', 'Kogi', 'Kwara', 'Nasarawa',
    'Niger', 'Plateau', 'Sokoto', 'Taraba', 'Yobe', 'Zamfara'
}

def sf(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except: return default

def main():
    wb = load_workbook(INPUT_FILE)
    ws = wb['LGA_DATA']
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    num_cols = len(headers)

    def ci(name):
        try: return headers.index(name)
        except: return None

    data = []
    for r in range(3, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, num_cols + 1)]
        if row[0] is not None:
            data.append(row)

    STATE = ci('State')
    LGA = ci('LGA Name')
    CONFLICT = ci('Conflict History')
    LITERACY = ci('Adult Literacy Rate Pct')
    MALE_LIT = ci('Male Literacy Rate Pct')
    FEMALE_LIT = ci('Female Literacy Rate Pct')
    URBAN = ci('Urban Pct')
    PRIMARY_ENROLL = ci('Primary Enrollment Pct')
    SECONDARY_ENROLL = ci('Secondary Enrollment Pct')
    GENDER_PARITY = ci('Gender Parity Index')
    OUT_OF_SCHOOL = ci('Out of School Children Pct')

    # ================================================================
    # FIX A: MODERATE CONFLICT CATEGORY
    # ================================================================
    print("--- FIX A: MODERATE Conflict Category ---")
    # Strategy: convert ~50% of SEVERE LGAs in peripheral conflict states to MODERATE
    # States where SEVERE should be partially downgraded:
    # - Katsina: 34 SEVERE -> ~15 MODERATE
    # - Kaduna: 23 SEVERE -> ~10 MODERATE
    # - Niger: 25 SEVERE -> ~10 MODERATE
    # - Sokoto: 23 SEVERE -> ~12 MODERATE (had non-aggression pact)
    # - Kebbi: 21 SEVERE -> ~8 MODERATE
    # Also convert some LOW_IMPACT in Bauchi/Jigawa to MODERATE
    # And some SPARED in Gombe/Nasarawa/Plateau to LOW_IMPACT

    moderate_rates = {
        'Katsina': 0.45,   # ~15 of 34
        'Kaduna': 0.40,    # ~9 of 23
        'Niger': 0.40,     # ~10 of 25
        'Sokoto': 0.52,    # ~12 of 23 (pact meant less fighting)
        'Kebbi': 0.38,     # ~8 of 21
        'FCT': 0.50,       # ~3 of 6 (quickly recaptured)
        'Adamawa': 0.33,   # ~7 of 21
        'Yobe': 0.30,      # ~5 of 17
    }

    low_impact_to_moderate = {
        'Bauchi': 0.40,    # ~8 of 20
        'Jigawa': 0.35,    # ~9 of 27
    }

    spared_to_low_impact = {
        'Gombe': 0.45,     # ~5 of 11
        'Nasarawa': 0.30,  # ~4 of 13
        'Plateau': 0.25,   # ~4 of 17
        'Taraba': 0.20,    # ~3 of 16
    }

    moderate_count = 0
    for r in data:
        state = r[STATE]
        conflict = r[CONFLICT]

        # SEVERE -> MODERATE in peripheral conflict states
        if conflict == 'SEVERE' and state in moderate_rates:
            if rng.random() < moderate_rates[state]:
                r[CONFLICT] = 'MODERATE'
                moderate_count += 1

        # LOW_IMPACT -> MODERATE in affected states
        elif conflict == 'LOW_IMPACT' and state in low_impact_to_moderate:
            if rng.random() < low_impact_to_moderate[state]:
                r[CONFLICT] = 'MODERATE'
                moderate_count += 1

        # SPARED -> LOW_IMPACT in spillover states
        elif conflict == 'SPARED' and state in spared_to_low_impact:
            if rng.random() < spared_to_low_impact[state]:
                r[CONFLICT] = 'LOW_IMPACT'

    cats = defaultdict(int)
    for r in data:
        cats[r[CONFLICT]] += 1
    print(f"  Added {moderate_count} MODERATE entries")
    print(f"  Categories: {dict(sorted(cats.items()))}")

    # ================================================================
    # FIX B: NORTHERN LITERACY BOOST
    # ================================================================
    print("\n--- FIX B: Northern Literacy Boost ---")
    # Current: 60.6%, target: 65-75%
    # Need ~+8pp average. Apply a modest uplift to non-conflict LGAs
    for r in data:
        if r[STATE] not in NORTHERN_STATES:
            continue

        conflict = r[CONFLICT]
        lit = sf(r[LITERACY])
        male_lit = sf(r[MALE_LIT])
        female_lit = sf(r[FEMALE_LIT])
        urban = sf(r[URBAN])

        if conflict in ('AL_SHAHID_CONTROL',):
            # No change for al-shahid zones (already calibrated)
            boost = 0
        elif conflict in ('SEVERE', 'OCCUPATION'):
            boost = 3 + rng.normal(0, 1)
        elif conflict == 'MODERATE':
            boost = 6 + rng.normal(0, 1.5)
        elif conflict == 'LOW_IMPACT':
            boost = 8 + rng.normal(0, 1.5)
        else:
            # SPARED - biggest boost from girls' education mandate, community colleges
            if urban > 60:
                boost = 12 + rng.normal(0, 2)
            elif urban > 30:
                boost = 10 + rng.normal(0, 2)
            else:
                boost = 8 + rng.normal(0, 2)

        new_lit = min(98, lit + boost)
        r[LITERACY] = round(new_lit, 1)

        # Proportional boost to male/female
        if lit > 0:
            ratio = new_lit / lit
        else:
            ratio = 1.0
        r[MALE_LIT] = round(min(99, male_lit * ratio), 1)
        r[FEMALE_LIT] = round(min(97, female_lit * ratio * 0.98), 1)

        # Ensure male >= female
        if sf(r[MALE_LIT]) < sf(r[FEMALE_LIT]):
            r[MALE_LIT] = round(sf(r[FEMALE_LIT]) + rng.uniform(0, 3), 1)

        # Update enrollment and out-of-school for consistency
        if PRIMARY_ENROLL:
            r[PRIMARY_ENROLL] = round(min(99, max(sf(r[PRIMARY_ENROLL]), new_lit * 1.05 + rng.normal(0, 2))), 1)
        if SECONDARY_ENROLL:
            r[SECONDARY_ENROLL] = round(min(95, max(sf(r[SECONDARY_ENROLL]), new_lit * 0.78 + rng.normal(0, 3))), 1)
        if OUT_OF_SCHOOL:
            r[OUT_OF_SCHOOL] = round(max(2, min(60, 100 - new_lit * 0.92 + rng.normal(0, 2))), 1)
        if GENDER_PARITY:
            gp = sf(r[FEMALE_LIT]) / max(1, sf(r[MALE_LIT]))
            r[GENDER_PARITY] = round(max(0.4, min(1.2, gp + rng.normal(0, 0.02))), 2)

    north_lit = [sf(r[LITERACY]) for r in data if r[STATE] in NORTHERN_STATES]
    print(f"  New north mean literacy: {np.mean(north_lit):.1f}%")

    # Write back
    print("\n--- Writing back ---")
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            ws.cell(row=r_idx + 3, column=c_idx + 1, value=val)

    wb.save(OUTPUT_FILE)
    print(f"Saved to {OUTPUT_FILE}")

    # Verification
    print("\n=== VERIFICATION ===")
    cats = defaultdict(int)
    for r in data:
        cats[r[CONFLICT]] += 1
    print(f"  Conflict: {dict(sorted(cats.items()))}")
    print(f"  North literacy: {np.mean(north_lit):.1f}%  (target: 65-75%)")

if __name__ == '__main__':
    main()
