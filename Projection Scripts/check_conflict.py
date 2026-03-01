#!/usr/bin/env python3
"""Quick check of conflict categories by state."""
from openpyxl import load_workbook
from collections import defaultdict

wb = load_workbook("nigeria_lga_polsim_2058.xlsx", data_only=True)
ws = wb['LGA_DATA']
headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

def ci(name):
    try: return headers.index(name)
    except: return None

STATE = ci('State')
CONFLICT = ci('Conflict History')
LGA = ci('LGA Name')

data = []
for r in range(3, ws.max_row + 1):
    row = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
    if row[0]: data.append(row)

state_conflict = defaultdict(lambda: defaultdict(int))
for r in data:
    state_conflict[r[STATE]][r[CONFLICT]] += 1

print("State -> Conflict Category breakdown:")
for state in sorted(state_conflict.keys()):
    cats = state_conflict[state]
    print(f"  {state:20s}: {dict(cats)}")
wb.close()
